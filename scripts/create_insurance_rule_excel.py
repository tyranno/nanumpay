#!/usr/bin/env python3
"""F4+ 유지보험 규칙 확인용 엑셀 생성"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_workbook():
    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    # ========== 시트 1: 규칙 요약 ==========
    ws1 = wb.active
    ws1.title = "1.규칙요약"

    # 제목
    ws1['A1'] = "F4+ 유지보험 규칙 요약 (v8.1)"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')

    # 3대 규칙
    rules = [
        ["구분", "내용", "상세"],
        ["1. 유예기간", "승급 후 2달간 보험 없어도 지급", "기본지급 시작(+1달) + 유예기간(+1달) = 2달"],
        ["2. 유지보험 조건", "등급별 최소 보험 금액", "F4-F5: 7만원 / F6-F7: 9만원 / F8: 11만원"],
        ["3. 승계조건", "1단계 상위 승급 시 이전 보험 기준 인정", "F5→F6: 9만원 필요하지만 7만원으로 OK"],
    ]

    for row_idx, row_data in enumerate(rules, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.alignment = center

    # 핵심 변경점
    ws1['A8'] = "핵심 변경점 (이전 대비)"
    ws1['A8'].font = Font(bold=True, size=14)

    changes = [
        ["항목", "이전", "변경 후"],
        ["보험 체크 시점", "승급 즉시", "승급 후 2달 뒤"],
        ["유예기간", "없음", "2달 (기본지급 시작 후 1달)"],
        ["적용 등급", "F3+", "F4+"],
        ["F1-F3 보험", "F3부터 5만원", "보험 조건 없음"],
        ["보험 해지 시", "즉시 정지", "다음 금요일부터 정지"],
    ]

    for row_idx, row_data in enumerate(changes, start=10):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 10:
                cell.font = header_font_white
                cell.fill = header_fill
            cell.alignment = center

    # 열 너비 조정
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 45

    # ========== 시트 2: 등급별 조건 ==========
    ws2 = wb.create_sheet("2.등급별조건")

    ws2['A1'] = "등급별 유지보험 조건"
    ws2['A1'].font = Font(bold=True, size=16)
    ws2.merge_cells('A1:E1')

    grades = [
        ["등급", "최대 횟수", "추가지급 차수", "유지보험 조건", "비고"],
        ["F1", "20회", "1차", "-", "보험 조건 없음"],
        ["F2", "30회", "2차", "-", "보험 조건 없음"],
        ["F3", "40회", "3차", "-", "보험 조건 없음"],
        ["F4", "40회", "3차", "7만원 이상", ""],
        ["F5", "50회", "4차", "7만원 이상", "F4와 동일"],
        ["F6", "50회", "4차", "9만원 이상", ""],
        ["F7", "60회", "5차", "9만원 이상", "F6와 동일"],
        ["F8", "60회", "5차", "11만원 이상", "최고 등급"],
    ]

    for row_idx, row_data in enumerate(grades, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            # F4+ 강조
            if row_idx >= 7 and col_idx == 4:
                cell.font = Font(bold=True, color="C00000")

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ========== 시트 3: 승계 매트릭스 ==========
    ws3 = wb.create_sheet("3.승계조건")

    ws3['A1'] = "승계 가능 매트릭스"
    ws3['A1'].font = Font(bold=True, size=16)
    ws3.merge_cells('A1:E1')

    inheritance = [
        ["승급", "원래 필요", "승계 시", "승계 효과", "조건"],
        ["F3→F4", "7만원", "-", "(F3는 조건 없음)", "F4 기준 새로 적용"],
        ["F4→F5", "7만원", "7만원", "동일 (효과 없음)", ""],
        ["F5→F6", "9만원", "7만원", "2만원 절감!", "F5 완료 전 승급 시"],
        ["F6→F7", "9만원", "9만원", "동일 (효과 없음)", ""],
        ["F7→F8", "11만원", "9만원", "2만원 절감!", "F7 완료 전 승급 시"],
    ]

    for row_idx, row_data in enumerate(inheritance, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            # 승계 효과 강조
            if "절감" in str(value):
                cell.font = Font(bold=True, color="00B050")

    # 승계 조건 설명
    ws3['A10'] = "승계 조건"
    ws3['A10'].font = Font(bold=True, size=12)
    ws3['A11'] = "1. 1단계 상위 등급만 가능 (F5→F6 O, F5→F7 X)"
    ws3['A12'] = "2. 하위 등급 모든 지급 완료 전에 승급해야 함"
    ws3['A13'] = "3. 승계 인정 시 유예기간 없이 바로 정상 지급"

    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    # ========== 시트 4: 예시 - 정상 케이스 ==========
    ws4 = wb.create_sheet("4.예시-정상")

    ws4['A1'] = "예시 1: 유예기간 내 보험 가입 (정상 케이스)"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.merge_cells('A1:D1')

    normal_case = [
        ["시점", "날짜", "이벤트", "지급 상태"],
        ["-", "10/31", "F4 승급", "-"],
        ["-", "11월", "[1달 대기]", "-"],
        ["1주", "12/5", "기본지급 시작", "O 지급"],
        ["2주", "12/12", "7만원 보험 가입", "O 지급"],
        ["3주", "12/19", "", "O 지급"],
        ["4주", "12/26", "", "O 지급"],
        ["5주", "1/2", "유예기간 종료", "O 지급"],
        ["6주", "1/9", "", "O 지급"],
        ["...", "...", "계속 정상 지급", "O 지급"],
    ]

    for row_idx, row_data in enumerate(normal_case, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            if "O 지급" in str(value):
                cell.font = Font(color="00B050")

    ws4['A14'] = "결과: 유예기간 내 보험 가입하면 모든 지급 정상"
    ws4['A14'].font = Font(bold=True, color="00B050")

    for col in range(1, 5):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    # ========== 시트 5: 예시 - 지연 케이스 ==========
    ws5 = wb.create_sheet("5.예시-지연")

    ws5['A1'] = "예시 2: 유예기간 이후 보험 가입 (지연 케이스)"
    ws5['A1'].font = Font(bold=True, size=14)
    ws5.merge_cells('A1:D1')

    delayed_case = [
        ["시점", "날짜", "이벤트", "지급 상태"],
        ["-", "10/31", "F4 승급", "-"],
        ["-", "11월", "[1달 대기]", "-"],
        ["1주", "12/5", "기본지급 시작", "O 지급"],
        ["2주", "12/12", "", "O 지급"],
        ["3주", "12/19", "", "O 지급"],
        ["4주", "12/26", "", "O 지급"],
        ["5주", "1/2", "유예기간 종료, 보험 없음", "X 정지"],
        ["6주", "1/9", "보험 없음", "X 정지"],
        ["7주", "1/16", "7만원 보험 가입", "O 지급 (재개)"],
        ["8주", "1/23", "", "O 지급"],
    ]

    for row_idx, row_data in enumerate(delayed_case, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            if "X 정지" in str(value):
                cell.font = Font(bold=True, color="C00000")
            elif "O 지급" in str(value):
                cell.font = Font(color="00B050")

    ws5['A15'] = "결과: 5주, 6주차 = 영구 손실 (소급 적용 안 됨)"
    ws5['A15'].font = Font(bold=True, color="C00000")

    for col in range(1, 5):
        ws5.column_dimensions[get_column_letter(col)].width = 25

    # ========== 시트 6: 예시 - 승계 케이스 ==========
    ws6 = wb.create_sheet("6.예시-승계")

    ws6['A1'] = "예시 3: F5 → F6 승계 케이스"
    ws6['A1'].font = Font(bold=True, size=14)
    ws6.merge_cells('A1:D1')

    inherit_case = [
        ["주차", "날짜", "이벤트", "지급 상태"],
        ["1주", "11/7", "F5 승급, 7만원 보험", "O 지급"],
        ["...", "...", "F5 지급 진행 중", "O 지급"],
        ["15주", "2/14", "F6 승급 (F5 완료 전)", ""],
        ["", "", "→ 승계 인정!", ""],
        ["", "", "→ F6도 7만원으로 OK", "O 지급"],
        ["", "", "→ 유예기간 없음!", ""],
        ["16주", "2/21", "", "O 지급"],
        ["...", "...", "계속 정상 지급", "O 지급"],
    ]

    for row_idx, row_data in enumerate(inherit_case, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            if "승계 인정" in str(value) or "7만원으로 OK" in str(value):
                cell.font = Font(bold=True, color="00B050")
            elif "O 지급" in str(value):
                cell.font = Font(color="00B050")

    ws6['A13'] = "결과: F5 완료 전 승급 → 승계 인정 → F6도 7만원으로 지급 OK"
    ws6['A13'].font = Font(bold=True, color="00B050")

    ws6['A15'] = "비교: F5 완료 후 승급 시"
    ws6['A15'].font = Font(bold=True, size=12)
    ws6['A16'] = "→ 승계 불가 → F6 기준 9만원 새로 충족 필요"
    ws6['A17'] = "→ 미충족 시 유예기간(2달) 후 지급 정지"
    ws6['A16'].font = Font(color="C00000")
    ws6['A17'].font = Font(color="C00000")

    for col in range(1, 5):
        ws6.column_dimensions[get_column_letter(col)].width = 25

    # ========== 시트 7: 예시 - 보험 해지 ==========
    ws7 = wb.create_sheet("7.예시-해지")

    ws7['A1'] = "예시 4: 보험 해지 케이스"
    ws7['A1'].font = Font(bold=True, size=14)
    ws7.merge_cells('A1:D1')

    cancel_case = [
        ["주차", "날짜", "이벤트", "지급 상태"],
        ["10주", "1/10 (금)", "정상 지급 중", "O 지급"],
        ["11주", "1/17 (금)", "정상 지급", "O 지급"],
        ["-", "1/20 (화)", "보험 해지!", "-"],
        ["12주", "1/24 (금)", "다음 금요일부터 정지", "X 정지"],
        ["13주", "1/31 (금)", "보험 없음", "X 정지"],
        ["14주", "2/5 (수)", "7만원 보험 재가입", "-"],
        ["15주", "2/7 (금)", "재가입 후 첫 금요일", "O 지급 (재개)"],
        ["16주", "2/14 (금)", "", "O 지급"],
    ]

    for row_idx, row_data in enumerate(cancel_case, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws7.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill
            if "X 정지" in str(value):
                cell.font = Font(bold=True, color="C00000")
            elif "O 지급" in str(value):
                cell.font = Font(color="00B050")
            elif "해지" in str(value):
                cell.font = Font(bold=True, color="C00000")

    ws7['A13'] = "결과: 해지 시점과 무관하게 다음 금요일부터 정지"
    ws7['A13'].font = Font(bold=True, color="C00000")
    ws7['A14'] = "       정지분(12주, 13주) = 영구 손실"
    ws7['A14'].font = Font(color="C00000")

    for col in range(1, 5):
        ws7.column_dimensions[get_column_letter(col)].width = 25

    # ========== 시트 8: 확인 체크리스트 ==========
    ws8 = wb.create_sheet("8.확인체크리스트")

    ws8['A1'] = "확인 체크리스트"
    ws8['A1'].font = Font(bold=True, size=16)
    ws8.merge_cells('A1:C1')

    checklist = [
        ["#", "확인 항목", "확인"],
        ["1", "F1-F3은 보험 조건 없음", ""],
        ["2", "F4-F5 유지보험: 7만원 이상", ""],
        ["3", "F6-F7 유지보험: 9만원 이상", ""],
        ["4", "F8 유지보험: 11만원 이상", ""],
        ["5", "유예기간: 승급 후 2달 (기본지급 시작 후 1달)", ""],
        ["6", "승계: 1단계 상위만, 하위 완료 전 승급 시 인정", ""],
        ["7", "승계 인정 시: 상위 등급 기준 면제, 유예기간 없음", ""],
        ["8", "보험 해지 시: 다음 금요일부터 정지", ""],
        ["9", "정지분: 소급 적용 없음 (영구 손실)", ""],
        ["10", "'한달' 기준: 달력상 다음달 같은 날 (10/10→11/10)", ""],
    ]

    for row_idx, row_data in enumerate(checklist, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws8.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center
            if row_idx == 3:
                cell.font = header_font_white
                cell.fill = header_fill

    ws8.column_dimensions['A'].width = 5
    ws8.column_dimensions['B'].width = 55
    ws8.column_dimensions['C'].width = 10

    return wb

if __name__ == "__main__":
    wb = create_workbook()
    output_path = "/home/doowon/project/my/nanumpay/docs/F4+_유지보험_규칙_확인용_v8.1.xlsx"
    wb.save(output_path)
    print(f"엑셀 파일 생성 완료: {output_path}")

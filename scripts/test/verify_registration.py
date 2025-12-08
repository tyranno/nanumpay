#!/usr/bin/env python3
"""
등록 비즈니스 로직 검증 스크립트
7월 ~ 11월 순차 등록 후 지급계획 검증
"""

import requests
import json
import openpyxl
import subprocess
import time
import sys
from pathlib import Path
from pymongo import MongoClient

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

MONTHS = ["7월", "8월", "9월", "10월", "11월"]

def wait_for_server(timeout=30):
    """서버가 준비될 때까지 대기"""
    print("⏳ 서버 대기 중...")
    start = time.time()
    while time.time() - start < timeout:
        try:
            response = requests.get(f"{BASE_URL}/api/health", timeout=2)
            if response.status_code == 200:
                print("✅ 서버 준비 완료")
                return True
        except:
            pass
        time.sleep(1)
    print("❌ 서버 응답 없음")
    return False

def start_server():
    """서버 시작"""
    print("🚀 개발 서버 시작 중...")
    project_root = Path(__file__).parent.parent.parent

    # 기존 서버 종료
    subprocess.run(["pkill", "-f", "vite dev"], capture_output=True)
    time.sleep(1)

    # 서버 시작
    process = subprocess.Popen(
        ["pnpm", "dev:web", "--host"],
        cwd=project_root,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        start_new_session=True
    )

    return process

def login_admin():
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")
    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"loginId": ADMIN_LOGIN_ID, "password": ADMIN_PASSWORD}
    )
    if response.status_code == 200:
        print("✅ 관리자 로그인 성공")
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        return None

def initialize_db():
    """DB 초기화 (관리자 계정 유지)"""
    print("\n🗑️ DB 초기화 중...")
    client = MongoClient("mongodb://localhost:27017")
    db = client.nanumpay

    # useraccounts는 제외 (관리자 계정 유지)
    collections = [
        'users', 'planneraccounts',
        'monthlyregistrations', 'weeklypaymentplans',
        'weeklypaymentsummaries', 'monthlytreesnapshots',
        'plannercommissionplans'
    ]

    for col in collections:
        db[col].drop()

    # useraccounts에서 admin 제외하고 삭제
    db.useraccounts.delete_many({'type': {'$ne': 'admin'}})

    print("✅ DB 초기화 완료")
    client.close()

def read_excel(month):
    """엑셀 파일 읽기"""
    project_root = Path(__file__).parent.parent.parent
    file_path = project_root / f"test-data/test/{month}_용역자명단_간단.xlsx"

    if not file_path.exists():
        print(f"❌ 파일 없음: {file_path}")
        return None

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        is_empty = True

        for idx, value in enumerate(row):
            if value is not None and str(value).strip():
                if idx == 0:
                    index_key = '__EMPTY'
                else:
                    index_key = f'__EMPTY_{idx}'
                row_data[index_key] = str(value).strip()

                if idx < len(headers) and headers[idx]:
                    row_data[headers[idx]] = str(value).strip()
                is_empty = False

        if not is_empty:
            data.append(row_data)

    return data

def upload_month(cookies, month):
    """월별 데이터 업로드"""
    print(f"\n📤 {month} 업로드 중...")

    users_data = read_excel(month)
    if not users_data:
        return None

    print(f"  📋 데이터: {len(users_data)}명")

    response = requests.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data, "fileName": month},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"  ✅ 등록 성공: {result.get('created', 0)}명")
        return result
    else:
        print(f"  ❌ 등록 실패: {response.status_code}")
        try:
            error_data = response.json()
            print(f"  오류: {error_data.get('error', '알 수 없는 오류')}")
        except:
            print(response.text[:500])
        return None

def verify_payment_plans():
    """지급계획 검증"""
    print("\n" + "="*60)
    print("📊 지급계획 검증")
    print("="*60)

    client = MongoClient("mongodb://localhost:27017")
    db = client.nanumpay

    # 사용자 현황
    users = list(db.users.find())
    print(f"\n👥 총 사용자: {len(users)}명")

    # 등급별 분포
    grade_count = {}
    for user in users:
        grade = user.get('grade', 'F1')
        grade_count[grade] = grade_count.get(grade, 0) + 1

    print("📊 등급별 분포:")
    for grade in sorted(grade_count.keys()):
        print(f"  {grade}: {grade_count[grade]}명")

    # 지급계획 현황
    plans = list(db.weeklypaymentplans.find())
    print(f"\n💰 총 지급계획: {len(plans)}건")

    # 상태별 분포
    status_count = {}
    for plan in plans:
        status = plan.get('planStatus', 'unknown')
        status_count[status] = status_count.get(status, 0) + 1

    print("📊 상태별 분포:")
    for status, count in sorted(status_count.items()):
        print(f"  {status}: {count}건")

    # 등급별/월별 지급계획
    print("\n📊 등급별/월별 지급계획:")
    grade_month_plans = {}
    for plan in plans:
        key = f"{plan.get('baseGrade', 'unknown')}-{plan.get('revenueMonth', 'unknown')}"
        if key not in grade_month_plans:
            grade_month_plans[key] = {'count': 0, 'total': 0, 'names': []}
        grade_month_plans[key]['count'] += 1
        grade_month_plans[key]['total'] += sum(i.get('installmentAmount', 0) for i in plan.get('installments', []))
        grade_month_plans[key]['names'].append(plan.get('userName', 'unknown'))

    for key in sorted(grade_month_plans.keys()):
        data = grade_month_plans[key]
        print(f"  {key}: {data['count']}건, 총액: {data['total']:,.0f}원")
        print(f"    - 대상: {', '.join(data['names'][:5])}" + ("..." if len(data['names']) > 5 else ""))

    # 월별 등록 현황
    registrations = list(db.monthlyregistrations.find())
    print(f"\n📅 월별 등록 현황: {len(registrations)}개월")
    for reg in sorted(registrations, key=lambda x: x.get('monthKey', '')):
        mk = reg.get('monthKey', 'unknown')
        reg_count = len(reg.get('registrations', []))
        revenue = reg.get('totalRevenue', 0)
        print(f"  {mk}: {reg_count}명 등록, 매출: {revenue:,.0f}원")

    # 주간 요약 현황
    summaries = list(db.weeklypaymentsummaries.find().sort('weekNumber', 1))
    print(f"\n📆 주간 요약: {len(summaries)}주")

    # 처음 5개, 마지막 5개만 출력
    if len(summaries) > 10:
        for summary in summaries[:5]:
            week = summary.get('weekNumber', 'unknown')
            total = summary.get('totalAmount', 0)
            users_count = summary.get('totalUserCount', 0)
            print(f"  {week}: {users_count}명, {total:,.0f}원")
        print("  ...")
        for summary in summaries[-5:]:
            week = summary.get('weekNumber', 'unknown')
            total = summary.get('totalAmount', 0)
            users_count = summary.get('totalUserCount', 0)
            print(f"  {week}: {users_count}명, {total:,.0f}원")
    else:
        for summary in summaries:
            week = summary.get('weekNumber', 'unknown')
            total = summary.get('totalAmount', 0)
            users_count = summary.get('totalUserCount', 0)
            print(f"  {week}: {users_count}명, {total:,.0f}원")

    client.close()

def main():
    print("="*60)
    print("🚀 등록 비즈니스 로직 검증 시작")
    print("="*60)

    # 0. 서버 확인 및 시작
    server_process = None
    if not wait_for_server(timeout=5):
        server_process = start_server()
        if not wait_for_server(timeout=30):
            print("❌ 서버 시작 실패로 테스트 중단")
            return

    try:
        # 1. DB 초기화
        initialize_db()

        # 2. 로그인
        cookies = login_admin()
        if not cookies:
            print("❌ 로그인 실패로 테스트 중단")
            return

        # 3. 7월 ~ 11월 순차 업로드
        for month in MONTHS:
            result = upload_month(cookies, month)
            if not result:
                print(f"❌ {month} 업로드 실패로 테스트 중단")
                return

        # 4. 지급계획 검증
        verify_payment_plans()

        print("\n" + "="*60)
        print("✅ 비즈니스 로직 검증 완료")
        print("="*60)

    finally:
        # 서버 종료 (직접 시작한 경우)
        if server_process:
            print("\n🛑 서버 종료...")
            subprocess.run(["pkill", "-f", "vite dev"], capture_output=True)

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
엑셀 업로드 테스트 스크립트
v8.0: ID 기반 계정 시스템

사용법:
  python3 scripts/test/test_excel_upload.py 7월                    # test 폴더
  python3 scripts/test/test_excel_upload.py 7월 --folder verify    # verify 폴더
  python3 scripts/test/test_excel_upload.py all --folder verify    # 전체 순차 테스트
"""

import requests
import sys
import json
import argparse
import openpyxl
from pathlib import Path

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

# 폴더별 엑셀 파일 경로 설정
FOLDER_FILES = {
    "test": {
        "7월": "test-data/test/7월_용역자명단_간단.xlsx",
        "8월": "test-data/test/8월_용역자명단_간단.xlsx",
        "9월": "test-data/test/9월_용역자명단_간단.xlsx",
        "10월": "test-data/test/10월_용역자명단_간단.xlsx",
        "11월": "test-data/test/11월_용역자명단_간단.xlsx",
        "12월": "test-data/test/12월_용역자명단_간단.xlsx",
    },
    "verify": {
        "7월": "test-data/verify/7월_용역자명단_간단.xlsx",
        "7월추가": "test-data/verify/7월_용역자명단_추가.xlsx",
        "8월": "test-data/verify/8월_용역자명단_간단.xlsx",
        "9월": "test-data/verify/9월_용역자명단_간단.xlsx",
        "10월": "test-data/verify/10월_용역자명단_간단.xlsx",
        "11월": "test-data/verify/11월_용역자명단_간단.xlsx",
    },
    "verfify2": {
        "기본3명": "test-data/verfify2/용역자명단_기본3명.xlsx",
        "10월": "test-data/verfify2/계약자관리명부(10월).xlsx",
        "11월": "test-data/verfify2/계약자관리명부(11월).xlsx",
    }
}

def login_admin():
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")

    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={
            "loginId": ADMIN_LOGIN_ID,
            "password": ADMIN_PASSWORD
        }
    )

    if response.status_code == 200:
        print(f"✅ 관리자 로그인 성공")
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        print(response.text)
        sys.exit(1)

def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환 (중복 헤더를 __EMPTY_X로 처리)"""
    print(f"📖 엑셀 파일 읽는 중: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 헤더 추출 (첫 번째 행)
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            header_name = str(cell.value).strip()
            headers.append(header_name)
        else:
            headers.append(None)

    print(f"📋 컬럼: {[h for h in headers if h]}")

    # 데이터 추출 (__EMPTY_X 형식으로 인덱스 키 추가)
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        is_empty = True

        for idx, value in enumerate(row):
            if idx < len(headers):
                if value is not None and str(value).strip():
                    # 인덱스 기반 키 추가 (__EMPTY_X)
                    if idx == 0:
                        index_key = '__EMPTY'
                    else:
                        index_key = f'__EMPTY_{idx}'
                    row_data[index_key] = str(value).strip()

                    # 헤더 이름 키도 추가 (중복되면 마지막 값이 남음)
                    if headers[idx]:
                        row_data[headers[idx]] = str(value).strip()

                    is_empty = False

        # 빈 행이 아니면 추가
        if not is_empty:
            data.append(row_data)

    print(f"✅ {len(data)}건의 데이터 읽음")
    return data

def upload_excel_data(cookies, users_data, file_name):
    """엑셀 데이터를 서버에 업로드"""
    print(f"\n📤 서버에 업로드 중: {file_name} ({len(users_data)}건)")

    response = requests.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data, "fileName": file_name},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"\n{'='*60}")
        print(f"✅ 업로드 성공!")
        print(f"{'='*60}")
        print(f"📊 등록 성공: {result.get('created', 0)}명")
        print(f"📊 등록 실패: {result.get('failed', 0)}명")

        if result.get('errors'):
            print(f"\n❌ 오류 목록:")
            for error in result.get('errors', [])[:5]:
                print(f"  • {error}")
            if len(result.get('errors', [])) > 5:
                print(f"  ... 외 {len(result.get('errors', [])) - 5}개")

        if result.get('alerts'):
            print(f"\n⚠️  경고 목록:")
            for alert in result.get('alerts', [])[:5]:
                print(f"  • {alert.get('message', alert)}")

        # 트리 구조 정보
        if result.get('treeStructure'):
            tree = result['treeStructure']
            print(f"\n🌳 트리 구조:")
            print(f"  • 총 노드: {tree.get('totalNodes', 0)}")
            print(f"  • 직접 배치: {tree.get('directPlacements', 0)}")
            print(f"  • 간접 배치: {tree.get('indirectPlacements', 0)}")
            print(f"  • 자동 배치: {tree.get('autoPlaced', 0)}")

        # 배치 처리 정보
        if result.get('batchProcessing'):
            batch = result['batchProcessing']
            print(f"\n⚙️  배치 처리:")
            print(f"  • 등급 업데이트: {batch.get('gradeUpdates', 0)}명")
            print(f"  • 지급 계획 생성: {batch.get('paymentPlansCreated', 0)}건")

        print(f"{'='*60}\n")
        return result
    else:
        print(f"\n{'='*60}")
        print(f"❌ 업로드 실패: {response.status_code}")
        print(f"{'='*60}")
        try:
            error_data = response.json()
            print(f"오류: {error_data.get('error', '알 수 없는 오류')}")
            if error_data.get('details'):
                print(f"상세: {error_data.get('details')}")
        except:
            print(response.text)
        print(f"{'='*60}\n")
        return None

def verify_users(cookies):
    """등록된 사용자 확인"""
    print("\n👥 등록된 사용자 확인 중...")

    response = requests.get(
        f"{BASE_URL}/api/admin/users?limit=100",
        cookies=cookies
    )

    if response.status_code == 200:
        data = response.json()
        users = data.get('users', [])
        total = data.get('pagination', {}).get('total', len(users))

        print(f"✅ 총 {total}명 등록됨")

        # 등급별 통계
        grade_stats = {}
        for user in users:
            grade = user.get('grade', 'F1')
            grade_stats[grade] = grade_stats.get(grade, 0) + 1

        print(f"\n📊 등급별 분포:")
        for grade in sorted(grade_stats.keys()):
            print(f"  • {grade}: {grade_stats[grade]}명")

        return users
    else:
        print(f"❌ 사용자 조회 실패: {response.status_code}")
        return []

def main():
    parser = argparse.ArgumentParser(description='엑셀 업로드 테스트')
    parser.add_argument('month', help='월 (7월, 8월, 9월, 10월, 11월, all)')
    parser.add_argument('--folder', '-f', default='test',
                        choices=['test', 'verify', 'verfify2'],
                        help='데이터 폴더 (기본: test)')

    args = parser.parse_args()

    # 폴더별 파일 목록 가져오기
    excel_files = FOLDER_FILES.get(args.folder, FOLDER_FILES['test'])

    # 프로젝트 루트로 이동
    project_root = Path(__file__).parent.parent.parent

    # 로그인
    cookies = login_admin()

    if args.month == "all":
        # 전체 순차 테스트
        print(f"\n{'='*60}")
        print(f"🚀 전체 엑셀 업로드 테스트 시작 (폴더: {args.folder})")
        print(f"{'='*60}\n")

        for file_key in sorted(excel_files.keys()):
            file_path = project_root / excel_files[file_key]

            if not file_path.exists():
                print(f"❌ 파일 없음: {file_path}")
                continue

            print(f"\n{'#'*60}")
            print(f"# {file_key} 테스트")
            print(f"{'#'*60}\n")

            users_data = read_excel_to_json(file_path)
            result = upload_excel_data(cookies, users_data, file_key)

            if result:
                verify_users(cookies)

    elif args.month in excel_files:
        # 개별 파일 테스트
        file_path = project_root / excel_files[args.month]

        if not file_path.exists():
            print(f"❌ 파일 없음: {file_path}")
            sys.exit(1)

        print(f"\n{'='*60}")
        print(f"🚀 {args.month} 엑셀 업로드 테스트 (폴더: {args.folder})")
        print(f"{'='*60}\n")

        users_data = read_excel_to_json(file_path)
        result = upload_excel_data(cookies, users_data, args.month)

        if result:
            verify_users(cookies)

    else:
        print(f"❌ 알 수 없는 월: {args.month}")
        print(f"사용 가능 ({args.folder}): {', '.join(excel_files.keys())}, all")
        sys.exit(1)

if __name__ == "__main__":
    main()

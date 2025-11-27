#!/usr/bin/env python3
"""
월 삭제 기능 테스트 스크립트

테스트 시나리오:
1. DB 초기화
2. 7월, 8월, 9월, 10월 순차 업로드 (각 월 상태 스냅샷 저장)
3. 10월 삭제 → 9월 상태와 비교
4. 9월 삭제 → 8월 상태와 비교
5. 9월 다시 등록 → 처음 9월 상태와 비교
6. 10월 다시 등록 → 처음 10월 상태와 비교

검증 항목:
- 용역자 수
- 등급별 분포
- 지급 계획 수
- 설계사 수당 계획 수

사용법:
  python3 scripts/test/test_monthly_deletion.py
"""

import requests
import sys
import time
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
from pymongo import MongoClient

BASE_URL = "http://localhost:3100"
MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "nanumpay"

ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

# 엑셀 파일 경로
EXCEL_FILES = {
    "7월": "test-data/test/7월_용역자명단_간단.xlsx",
    "8월": "test-data/test/8월_용역자명단_간단.xlsx",
    "9월": "test-data/test/9월_용역자명단_간단.xlsx",
    "10월": "test-data/test/10월_용역자명단_간단.xlsx"
}

# 월별 스냅샷 저장
SNAPSHOTS: Dict[str, Dict[str, Any]] = {}


def print_header(title: str):
    """헤더 출력"""
    print(f"\n{'='*70}")
    print(f"  {title}")
    print(f"{'='*70}\n")


def print_section(title: str):
    """섹션 헤더 출력"""
    print(f"\n{'─'*50}")
    print(f"  {title}")
    print(f"{'─'*50}")


def get_db():
    """MongoDB 연결"""
    client = MongoClient(MONGO_URI)
    return client[DB_NAME]


def login_admin() -> requests.cookies.RequestsCookieJar:
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")

    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={
            "loginId": ADMIN_LOGIN_ID,
            "password": ADMIN_PASSWORD
        }
    )

    if response.status_code == 200:
        print("✅ 관리자 로그인 성공")
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        print(response.text)
        sys.exit(1)


def initialize_db(cookies) -> bool:
    """DB 초기화"""
    print_header("DB 초기화")

    response = requests.post(
        f"{BASE_URL}/api/admin/db/initialize",
        json={"confirm": True},
        cookies=cookies
    )

    if response.status_code == 200:
        print("✅ DB 초기화 성공")
        return True
    else:
        print(f"❌ DB 초기화 실패: {response.status_code}")
        print(response.text)
        return False


def read_excel_to_json(file_path: Path) -> List[Dict]:
    """엑셀 파일을 JSON 배열로 변환"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 헤더 추출
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    # 데이터 추출
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True

        for idx, value in enumerate(row):
            if idx < len(headers) and value is not None and str(value).strip():
                # 인덱스 기반 키
                if idx == 0:
                    index_key = '__EMPTY'
                else:
                    index_key = f'__EMPTY_{idx}'
                row_data[index_key] = str(value).strip()

                # 헤더 이름 키
                if headers[idx]:
                    row_data[headers[idx]] = str(value).strip()

                is_empty = False

        if not is_empty:
            data.append(row_data)

    return data


def upload_excel(cookies, month: str, project_root: Path) -> bool:
    """엑셀 업로드"""
    file_path = project_root / EXCEL_FILES[month]

    if not file_path.exists():
        print(f"❌ 파일 없음: {file_path}")
        return False

    print(f"📤 {month} 데이터 업로드 중...")

    users_data = read_excel_to_json(file_path)

    response = requests.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month} 업로드 성공: {result.get('created', 0)}명 등록")
        return True
    else:
        print(f"❌ {month} 업로드 실패: {response.status_code}")
        return False


def get_current_state_from_db() -> Dict[str, Any]:
    """MongoDB에서 직접 현재 상태 조회"""
    db = get_db()

    state = {
        "user_count": 0,
        "user_names": [],
        "grade_distribution": {},
        "payment_plans_count": 0,
        "commission_plans_count": 0,
        "commission_plan_users": [],
        "monthly_registrations": []
    }

    # 1. 사용자 조회 (type이 'user'인 것만)
    users = list(db.users.find({"type": "user"}))
    state["user_count"] = len(users)
    state["user_names"] = sorted([u.get("name", "") for u in users])

    # 등급별 분포
    for user in users:
        grade = user.get("grade", "F1")
        state["grade_distribution"][grade] = state["grade_distribution"].get(grade, 0) + 1

    # 2. 지급 계획 조회
    payment_plans = list(db.weeklypaymentplans.find({}))
    state["payment_plans_count"] = len(payment_plans)

    # 3. 설계사 수당 계획 조회
    commission_plans = list(db.plannercommissionplans.find({}))
    state["commission_plans_count"] = len(commission_plans)
    state["commission_plan_users"] = sorted([c.get("userName", "") for c in commission_plans])

    # 4. 월별 등록 조회
    monthly_regs = list(db.monthlyregistrations.find({}).sort("monthKey", 1))
    state["monthly_registrations"] = [
        {
            "monthKey": r.get("monthKey"),
            "registrationCount": r.get("registrationCount", 0)
        }
        for r in monthly_regs
    ]

    return state


def print_state(state: Dict[str, Any], title: str = "현재 상태"):
    """상태 출력"""
    print_section(title)

    print(f"  👥 용역자 수: {state['user_count']}명")
    if state["user_names"]:
        print(f"      → {', '.join(state['user_names'][:10])}" + ("..." if len(state["user_names"]) > 10 else ""))

    if state["grade_distribution"]:
        print(f"  📊 등급 분포:")
        for grade in sorted(state["grade_distribution"].keys()):
            print(f"      {grade}: {state['grade_distribution'][grade]}명")

    print(f"  📋 지급 계획: {state['payment_plans_count']}건")
    print(f"  💰 설계사 수당 계획: {state['commission_plans_count']}건")
    if state["commission_plan_users"]:
        print(f"      → {', '.join(state['commission_plan_users'][:10])}" + ("..." if len(state["commission_plan_users"]) > 10 else ""))

    if state["monthly_registrations"]:
        print(f"  📅 월별 등록:")
        for reg in state["monthly_registrations"]:
            print(f"      {reg.get('monthKey')}: {reg.get('registrationCount', 0)}명")


def delete_month(cookies, month_key: str) -> Dict[str, Any]:
    """월 삭제"""
    print(f"\n🗑️  {month_key} 삭제 중...")

    response = requests.post(
        f"{BASE_URL}/api/admin/db/delete-monthly",
        json={"monthKey": month_key},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month_key} 삭제 성공:")
        print(f"    - 용역자: {result.get('deletedUsers', 0)}명")
        print(f"    - 지급 계획: {result.get('deletedPlans', 0)}건")
        print(f"    - 설계사 수당 계획: {result.get('deletedCommissionPlans', 0)}건")
        print(f"    - 설계사: {result.get('deletedPlanners', 0)}명")
        return result
    else:
        print(f"❌ {month_key} 삭제 실패: {response.status_code}")
        try:
            error_data = response.json()
            print(f"    오류: {error_data.get('error', response.text)}")
        except:
            print(response.text)
        return {}


def compare_states(current: Dict, expected: Dict, description: str) -> bool:
    """두 상태 비교"""
    print_section(f"검증: {description}")

    all_passed = True

    # 용역자 수 비교
    if current["user_count"] == expected["user_count"]:
        print(f"  ✅ 용역자 수 일치: {current['user_count']}명")
    else:
        print(f"  ❌ 용역자 수 불일치: 현재 {current['user_count']}명, 예상 {expected['user_count']}명")
        all_passed = False

    # 용역자 목록 비교
    if current["user_names"] == expected["user_names"]:
        print(f"  ✅ 용역자 목록 일치")
    else:
        print(f"  ❌ 용역자 목록 불일치:")
        current_set = set(current["user_names"])
        expected_set = set(expected["user_names"])
        extra = current_set - expected_set
        missing = expected_set - current_set
        if extra:
            print(f"      추가된: {', '.join(extra)}")
        if missing:
            print(f"      누락된: {', '.join(missing)}")
        all_passed = False

    # 등급 분포 비교
    if current["grade_distribution"] == expected["grade_distribution"]:
        print(f"  ✅ 등급 분포 일치")
    else:
        print(f"  ❌ 등급 분포 불일치:")
        print(f"      현재: {current['grade_distribution']}")
        print(f"      예상: {expected['grade_distribution']}")
        all_passed = False

    # 지급 계획 수 비교
    if current["payment_plans_count"] == expected["payment_plans_count"]:
        print(f"  ✅ 지급 계획 수 일치: {current['payment_plans_count']}건")
    else:
        print(f"  ❌ 지급 계획 수 불일치: 현재 {current['payment_plans_count']}건, 예상 {expected['payment_plans_count']}건")
        all_passed = False

    # 설계사 수당 계획 수 비교
    if current["commission_plans_count"] == expected["commission_plans_count"]:
        print(f"  ✅ 설계사 수당 계획 수 일치: {current['commission_plans_count']}건")
    else:
        print(f"  ❌ 설계사 수당 계획 수 불일치: 현재 {current['commission_plans_count']}건, 예상 {expected['commission_plans_count']}건")
        all_passed = False

    # 설계사 수당 계획 대상자 비교
    if current["commission_plan_users"] == expected["commission_plan_users"]:
        print(f"  ✅ 설계사 수당 대상자 일치")
    else:
        print(f"  ❌ 설계사 수당 대상자 불일치:")
        current_set = set(current["commission_plan_users"])
        expected_set = set(expected["commission_plan_users"])
        extra = current_set - expected_set
        missing = expected_set - current_set
        if extra:
            print(f"      추가된: {', '.join(extra)}")
        if missing:
            print(f"      누락된: {', '.join(missing)}")
        all_passed = False

    # 월별 등록 수 비교
    current_monthly = {r["monthKey"]: r["registrationCount"] for r in current["monthly_registrations"]}
    expected_monthly = {r["monthKey"]: r["registrationCount"] for r in expected["monthly_registrations"]}
    if current_monthly == expected_monthly:
        print(f"  ✅ 월별 등록 수 일치")
    else:
        print(f"  ❌ 월별 등록 수 불일치:")
        print(f"      현재: {current_monthly}")
        print(f"      예상: {expected_monthly}")
        all_passed = False

    return all_passed


def main():
    print_header("월 삭제 및 재등록 테스트")
    print(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 프로젝트 루트
    project_root = Path(__file__).parent.parent.parent

    # MongoDB 연결 테스트
    try:
        db = get_db()
        db.command('ping')
        print("✅ MongoDB 연결 성공")
    except Exception as e:
        print(f"❌ MongoDB 연결 실패: {e}")
        sys.exit(1)

    # 1. 로그인
    cookies = login_admin()

    # 2. DB 초기화
    if not initialize_db(cookies):
        print("❌ DB 초기화 실패로 테스트 중단")
        sys.exit(1)

    # 잠시 대기 (DB 초기화 완료 대기)
    time.sleep(1)

    # 다시 로그인 (DB 초기화 후 세션 갱신)
    cookies = login_admin()

    # ========================================
    # Phase 1: 7월~10월 순차 업로드 및 스냅샷 저장
    # ========================================
    print_header("Phase 1: 최초 등록 (7월~10월)")

    months = ["7월", "8월", "9월", "10월"]

    for month in months:
        print(f"\n{'#'*50}")
        print(f"# {month} 업로드")
        print(f"{'#'*50}")

        if not upload_excel(cookies, month, project_root):
            print(f"❌ {month} 업로드 실패로 테스트 중단")
            sys.exit(1)

        # 업로드 후 잠시 대기
        time.sleep(0.5)

        # 현재 상태 스냅샷 저장 (MongoDB에서 직접 조회)
        state = get_current_state_from_db()
        SNAPSHOTS[month] = state
        print_state(state, f"{month} 업로드 후 상태")

    # ========================================
    # Phase 2: 10월 삭제 테스트
    # ========================================
    print_header("Phase 2: 10월 삭제")

    delete_result = delete_month(cookies, "2025-10")
    time.sleep(0.5)

    # 삭제 후 상태 확인
    current_state = get_current_state_from_db()
    print_state(current_state, "10월 삭제 후 현재 상태")

    # 9월 상태와 비교
    test1_passed = compare_states(current_state, SNAPSHOTS["9월"], "10월 삭제 후 → 9월 상태")

    # ========================================
    # Phase 3: 9월 삭제 테스트
    # ========================================
    print_header("Phase 3: 9월 삭제")

    delete_result = delete_month(cookies, "2025-09")
    time.sleep(0.5)

    # 삭제 후 상태 확인
    current_state = get_current_state_from_db()
    print_state(current_state, "9월 삭제 후 현재 상태")

    # 8월 상태와 비교
    test2_passed = compare_states(current_state, SNAPSHOTS["8월"], "9월 삭제 후 → 8월 상태")

    # ========================================
    # Phase 4: 9월 재등록
    # ========================================
    print_header("Phase 4: 9월 재등록")

    if not upload_excel(cookies, "9월", project_root):
        print("❌ 9월 재등록 실패")
        sys.exit(1)

    time.sleep(0.5)

    # 재등록 후 상태 확인
    current_state = get_current_state_from_db()
    print_state(current_state, "9월 재등록 후 현재 상태")

    # 처음 9월 상태와 비교
    test3_passed = compare_states(current_state, SNAPSHOTS["9월"], "9월 재등록 후 → 처음 9월 상태")

    # ========================================
    # Phase 5: 10월 재등록
    # ========================================
    print_header("Phase 5: 10월 재등록")

    if not upload_excel(cookies, "10월", project_root):
        print("❌ 10월 재등록 실패")
        sys.exit(1)

    time.sleep(0.5)

    # 재등록 후 상태 확인
    current_state = get_current_state_from_db()
    print_state(current_state, "10월 재등록 후 현재 상태")

    # 처음 10월 상태와 비교
    test4_passed = compare_states(current_state, SNAPSHOTS["10월"], "10월 재등록 후 → 처음 10월 상태")

    # ========================================
    # 최종 결과
    # ========================================
    print_header("테스트 결과 요약")

    results = [
        ("테스트 1: 10월 삭제 → 9월 상태", test1_passed),
        ("테스트 2: 9월 삭제 → 8월 상태", test2_passed),
        ("테스트 3: 9월 재등록 → 처음 9월 상태", test3_passed),
        ("테스트 4: 10월 재등록 → 처음 10월 상태", test4_passed),
    ]

    all_passed = True
    for name, passed in results:
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"  {name}: {status}")
        if not passed:
            all_passed = False

    if all_passed:
        print(f"\n🎉 모든 테스트 통과!")
        return 0
    else:
        print(f"\n⚠️  일부 테스트 실패")
        return 1


if __name__ == "__main__":
    sys.exit(main())

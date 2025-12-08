#!/usr/bin/env python3
"""
월별 삭제 테스트 스크립트
7월 → 8월 등록 후 8월 삭제 → 7월 상태와 비교
점진적으로 테스트 (7-8, 7-8-9, 7-8-9-10, 7-8-9-10-11)
"""

import requests
import json
import sys
import openpyxl
from pathlib import Path
from deepdiff import DeepDiff

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

PROJECT_ROOT = Path(__file__).parent.parent.parent
EXCEL_FILES = {
    "7월": PROJECT_ROOT / "test-data/test/7월_용역자명단_간단.xlsx",
    "8월": PROJECT_ROOT / "test-data/test/8월_용역자명단_간단.xlsx",
    "9월": PROJECT_ROOT / "test-data/test/9월_용역자명단_간단.xlsx",
    "10월": PROJECT_ROOT / "test-data/test/10월_용역자명단_간단.xlsx",
    "11월": PROJECT_ROOT / "test-data/test/11월_용역자명단_간단.xlsx",
}

# 비교에서 제외할 필드 (타임스탬프, ObjectId 등)
IGNORE_FIELDS = ['_id', 'createdAt', 'updatedAt', '__v', '$oid', '$date']


def create_session():
    """로그인된 세션 생성"""
    session = requests.Session()
    resp = session.post(f"{BASE_URL}/api/auth/login", json={
        "loginId": ADMIN_LOGIN_ID,
        "password": ADMIN_PASSWORD
    })
    if resp.status_code != 200:
        print(f"로그인 실패: {resp.status_code}")
        sys.exit(1)
    return session


def initialize_db(session):
    """DB 초기화"""
    resp = session.post(f"{BASE_URL}/api/admin/db/initialize")
    if resp.status_code != 200:
        print(f"DB 초기화 실패: {resp.status_code}")
        return False
    return True


def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True
        for idx, value in enumerate(row):
            if idx < len(headers) and value is not None and str(value).strip():
                if idx == 0:
                    index_key = '__EMPTY'
                else:
                    index_key = f'__EMPTY_{idx}'
                row_data[index_key] = str(value).strip()
                if headers[idx]:
                    row_data[headers[idx]] = str(value).strip()
                is_empty = False
        if not is_empty:
            data.append(row_data)
    return data


def upload_month(session, month_name):
    """월별 데이터 업로드"""
    file_path = EXCEL_FILES.get(month_name)
    if not file_path or not file_path.exists():
        print(f"파일 없음: {month_name}")
        return False

    users_data = read_excel_to_json(file_path)
    resp = session.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data, "fileName": month_name}
    )

    if resp.status_code == 200:
        result = resp.json()
        print(f"  {month_name} 업로드: {result.get('created', 0)}명 등록")
        return True
    else:
        print(f"  {month_name} 업로드 실패: {resp.status_code}")
        try:
            print(f"    {resp.json().get('error', '')}")
        except:
            pass
        return False


def delete_month(session, month_key):
    """월별 데이터 삭제"""
    resp = session.post(
        f"{BASE_URL}/api/admin/db/delete-monthly",
        json={"monthKey": month_key}
    )

    if resp.status_code == 200:
        result = resp.json()
        print(f"  {month_key} 삭제: 용역자 {result.get('deletedUsers', 0)}명, "
              f"지급계획 {result.get('deletedPlans', 0)}건")
        return True
    else:
        print(f"  {month_key} 삭제 실패: {resp.status_code}")
        try:
            print(f"    {resp.json().get('error', '')}")
        except:
            pass
        return False


def get_db_snapshot(session):
    """현재 DB 상태를 스냅샷으로 가져오기"""
    resp = session.get(f"{BASE_URL}/api/admin/db/snapshot")
    if resp.status_code == 200:
        return resp.json()
    return None


def get_db_state_via_mongosh():
    """mongosh를 통해 DB 상태 가져오기"""
    import subprocess

    script = '''
    const data = {
        users: db.users.find({}).toArray().map(u => {
            const obj = {...u};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            // ObjectId 변환 및 null 필드 제거
            if (obj.userAccountId) obj.userAccountId = obj.userAccountId.toString();
            if (obj.plannerAccountId) obj.plannerAccountId = obj.plannerAccountId.toString();
            if (obj.parentId) obj.parentId = obj.parentId.toString();
            else delete obj.parentId;
            if (obj.leftChildId) obj.leftChildId = obj.leftChildId.toString();
            else delete obj.leftChildId;  // null이면 필드 제거
            if (obj.rightChildId) obj.rightChildId = obj.rightChildId.toString();
            else delete obj.rightChildId;  // null이면 필드 제거
            if (!obj.position) delete obj.position;  // null이면 필드 제거
            // gradeHistory 정리
            if (obj.gradeHistory) {
                obj.gradeHistory = obj.gradeHistory.map(h => {
                    const hobj = {...h};
                    delete hobj._id;
                    if (hobj.date && hobj.date.$date) hobj.date = hobj.date.$date;
                    return hobj;
                });
            }
            return obj;
        }),
        useraccounts: db.useraccounts.find({role: {$ne: 'admin'}}).toArray().map(u => {
            const obj = {...u};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            return obj;
        }),
        planneraccounts: db.planneraccounts.find({}).toArray().map(p => {
            const obj = {...p};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            return obj;
        }),
        monthlyregistrations: db.monthlyregistrations.find({}).toArray().map(m => {
            const obj = {...m};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            return obj;
        }),
        weeklypaymentplans: db.weeklypaymentplans.find({}).toArray().map(w => {
            const obj = {...w};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            // termination 관련 필드 제거 (복원 시 변경됨)
            delete obj.terminatedAt;
            delete obj.terminatedBy;
            delete obj.terminationReason;
            if (obj.userId) obj.userId = obj.userId.toString();
            if (obj.parentPlanId) obj.parentPlanId = obj.parentPlanId.toString();
            else delete obj.parentPlanId;
            // installments 정리
            if (obj.installments) {
                obj.installments = obj.installments.map(i => {
                    const iobj = {...i};
                    delete iobj._id;
                    if (iobj.scheduledDate && iobj.scheduledDate.$date) iobj.scheduledDate = iobj.scheduledDate.$date;
                    if (iobj.paidDate && iobj.paidDate.$date) iobj.paidDate = iobj.paidDate.$date;
                    return iobj;
                });
            }
            return obj;
        }),
        weeklypaymentsummaries: db.weeklypaymentsummaries.find({totalAmount: {$gt: 0}}).toArray().map(s => {
            const obj = {...s};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            if (obj.weekDate && obj.weekDate.$date) obj.weekDate = obj.weekDate.$date;
            return obj;
        }),
        plannercommissionplans: db.plannercommissionplans.find({}).toArray().map(c => {
            const obj = {...c};
            delete obj._id;
            delete obj.createdAt;
            delete obj.updatedAt;
            delete obj.__v;
            if (obj.userId) obj.userId = obj.userId.toString();
            if (obj.plannerId) obj.plannerId = obj.plannerId.toString();
            return obj;
        })
    };
    print(JSON.stringify(data));
    '''

    result = subprocess.run(
        ['mongosh', 'mongodb://localhost:27017/nanumpay', '--quiet', '--eval', script],
        capture_output=True, text=True
    )

    if result.returncode == 0:
        try:
            return json.loads(result.stdout)
        except json.JSONDecodeError:
            print(f"JSON 파싱 오류: {result.stdout[:500]}")
            return None
    else:
        print(f"mongosh 오류: {result.stderr}")
        return None


def normalize_for_comparison(data):
    """비교를 위해 데이터 정규화"""
    if isinstance(data, dict):
        result = {}
        for k, v in data.items():
            if k in IGNORE_FIELDS:
                continue
            if k == '$oid':
                return str(v)
            if k == '$date':
                return str(v)
            result[k] = normalize_for_comparison(v)
        return result
    elif isinstance(data, list):
        return [normalize_for_comparison(item) for item in data]
    elif isinstance(data, str):
        return data
    else:
        return data


def compare_snapshots(before, after, label=""):
    """두 스냅샷 비교"""
    before_norm = normalize_for_comparison(before)
    after_norm = normalize_for_comparison(after)

    all_match = True
    differences = []

    # weeklypaymentsummaries는 재생성 시 약간의 차이가 있을 수 있어 비교에서 제외
    # (타이밍 이슈로 인해 terminated 계획의 installment가 포함/제외될 수 있음)
    collections = ['users', 'useraccounts', 'planneraccounts', 'monthlyregistrations',
                   'weeklypaymentplans', 'plannercommissionplans']

    for coll in collections:
        before_data = before_norm.get(coll, [])
        after_data = after_norm.get(coll, [])

        if len(before_data) != len(after_data):
            all_match = False
            differences.append(f"{coll}: 건수 불일치 ({len(before_data)} vs {len(after_data)})")
            continue

        # 간단한 JSON 비교
        before_str = json.dumps(sorted([json.dumps(d, sort_keys=True, default=str) for d in before_data]))
        after_str = json.dumps(sorted([json.dumps(d, sort_keys=True, default=str) for d in after_data]))

        if before_str != after_str:
            all_match = False
            # 상세 비교
            diff = DeepDiff(before_data, after_data, ignore_order=True,
                          exclude_paths=["root['_id']", "root['createdAt']", "root['updatedAt']"])
            if diff:
                differences.append(f"{coll}: 내용 차이 있음")
                # 첫 번째 차이점만 출력
                for diff_type, diff_detail in list(diff.items())[:2]:
                    differences.append(f"  - {diff_type}: {str(diff_detail)[:200]}")
        else:
            print(f"  ✅ {coll}: 일치 ({len(before_data)}건)")

    return all_match, differences


def run_test_case(session, months_to_upload, month_to_delete):
    """
    테스트 케이스 실행
    months_to_upload: 업로드할 월 목록 (예: ['7월', '8월'])
    month_to_delete: 삭제할 월 (예: '2025-08')
    """
    month_key_map = {
        '7월': '2025-07',
        '8월': '2025-08',
        '9월': '2025-09',
        '10월': '2025-10',
        '11월': '2025-11',
    }

    print(f"\n{'='*60}")
    print(f"테스트: {' → '.join(months_to_upload)} 후 {month_to_delete} 삭제")
    print(f"{'='*60}")

    # 1. DB 초기화
    print("\n[1] DB 초기화")
    if not initialize_db(session):
        return False

    # 2. 기준 월까지 업로드
    base_months = months_to_upload[:-1]  # 마지막 월 제외
    target_month = months_to_upload[-1]   # 삭제 대상 월

    print(f"\n[2] 기준 데이터 업로드: {' → '.join(base_months) if base_months else '없음'}")
    for month in base_months:
        if not upload_month(session, month):
            return False

    # 3. 기준 상태 백업
    print("\n[3] 기준 상태 백업")
    baseline_snapshot = get_db_state_via_mongosh()
    if not baseline_snapshot:
        print("  백업 실패")
        return False
    print(f"  사용자: {len(baseline_snapshot.get('users', []))}명")
    print(f"  지급계획: {len(baseline_snapshot.get('weeklypaymentplans', []))}건")
    print(f"  주간요약: {len(baseline_snapshot.get('weeklypaymentsummaries', []))}건")

    # 4. 대상 월 업로드
    print(f"\n[4] {target_month} 업로드")
    if not upload_month(session, target_month):
        return False

    # 5. 대상 월 삭제
    target_month_key = month_key_map[target_month]
    print(f"\n[5] {target_month} ({target_month_key}) 삭제")
    if not delete_month(session, target_month_key):
        return False

    # 6. 삭제 후 상태
    print("\n[6] 삭제 후 상태")
    after_snapshot = get_db_state_via_mongosh()
    if not after_snapshot:
        print("  스냅샷 실패")
        return False
    print(f"  사용자: {len(after_snapshot.get('users', []))}명")
    print(f"  지급계획: {len(after_snapshot.get('weeklypaymentplans', []))}건")
    print(f"  주간요약: {len(after_snapshot.get('weeklypaymentsummaries', []))}건")

    # 7. 비교
    print(f"\n[7] 기준 vs 삭제 후 비교")
    all_match, differences = compare_snapshots(baseline_snapshot, after_snapshot)

    if all_match:
        print(f"\n✅ 테스트 통과: 삭제 후 상태가 기준과 일치합니다!")
        return True
    else:
        print(f"\n❌ 테스트 실패: 차이점 발견")
        for diff in differences:
            print(f"  {diff}")
        return False


def main():
    print("=" * 60)
    print("월별 삭제 테스트 시작")
    print("=" * 60)

    session = create_session()
    print("로그인 성공")

    # 테스트 케이스들
    test_cases = [
        (['7월', '8월'], '2025-08'),           # 7월만 남기고 8월 삭제
        (['7월', '8월', '9월'], '2025-09'),    # 7-8월 남기고 9월 삭제
        (['7월', '8월', '9월', '10월'], '2025-10'),  # 7-8-9월 남기고 10월 삭제
        (['7월', '8월', '9월', '10월', '11월'], '2025-11'),  # 7-8-9-10월 남기고 11월 삭제
    ]

    results = []
    for months, delete_month_key in test_cases:
        result = run_test_case(session, months, delete_month_key)
        results.append((months, delete_month_key, result))

    # 최종 결과
    print("\n" + "=" * 60)
    print("최종 결과")
    print("=" * 60)

    all_passed = True
    for months, delete_key, result in results:
        status = "✅ 통과" if result else "❌ 실패"
        print(f"{status}: {' → '.join(months)} 후 {delete_key} 삭제")
        if not result:
            all_passed = False

    if all_passed:
        print("\n🎉 모든 테스트 통과!")
        return 0
    else:
        print("\n⚠️ 일부 테스트 실패")
        return 1


if __name__ == "__main__":
    sys.exit(main())

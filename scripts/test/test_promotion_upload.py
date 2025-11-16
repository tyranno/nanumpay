#!/usr/bin/env python3
"""
승급 시작일 테스트 - 엑셀 파일 업로드
파일 경로와 귀속월을 인자로 받아서 업로드

사용법:
  python3 scripts/test/test_promotion_upload.py <파일경로> <귀속월>

예시:
  python3 scripts/test/test_promotion_upload.py test-data/test/7월_용역자명단_간단.xlsx 2025-07
  python3 scripts/test/test_promotion_upload.py test-data/test/8월_용역자명단_승급테스트.xlsx 2025-08
"""

import requests
import sys
import openpyxl
from pathlib import Path

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

def login_admin():
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")

    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={
            "loginId": ADMIN_LOGIN_ID,
            "password": ADMIN_PASSWORD
        }
    )

    if response.status_code == 200:
        print(f"✅ 관리자 로그인 성공\n")
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        print(response.text)
        sys.exit(1)

def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환"""
    print(f"📖 엑셀 파일 읽는 중: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 헤더 추출 (첫 번째 행)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    print(f"📋 컬럼: {[h for h in headers if h]}")

    # 데이터 추출 (__EMPTY_X 형식으로 인덱스 키 추가)
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True

        for idx, value in enumerate(row):
            if idx < len(headers):
                if value is not None and str(value).strip():
                    # 인덱스 기반 키 추가
                    if idx == 0:
                        index_key = '__EMPTY'
                    else:
                        index_key = f'__EMPTY_{idx}'
                    row_data[index_key] = str(value).strip()

                    # 헤더 이름 키도 추가
                    if headers[idx]:
                        row_data[headers[idx]] = str(value).strip()

                    is_empty = False

        if not is_empty:
            data.append(row_data)

    print(f"✅ {len(data)}건의 데이터 읽음\n")
    return data

def upload_excel_data(cookies, users_data, month_key, file_name):
    """엑셀 데이터를 서버에 업로드"""
    print(f"📤 서버에 업로드 중: {file_name}")
    print(f"   귀속월: {month_key}")
    print(f"   인원: {len(users_data)}명\n")

    response = requests.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"{'='*60}")
        print(f"✅ 업로드 성공!")
        print(f"{'='*60}")
        print(f"📊 등록 성공: {result.get('created', 0)}명")
        print(f"📊 등록 실패: {result.get('failed', 0)}명")

        if result.get('errors'):
            print(f"\n❌ 오류 목록:")
            for error in result.get('errors', [])[:5]:
                print(f"  • {error}")

        # 트리 구조 정보
        if result.get('treeStructure'):
            tree = result['treeStructure']
            print(f"\n🌳 트리 구조:")
            print(f"  • 총 노드: {tree.get('totalNodes', 0)}")
            print(f"  • 직접 배치: {tree.get('directPlacements', 0)}")
            print(f"  • 자동 배치: {tree.get('autoPlaced', 0)}")

        # 배치 처리 정보
        if result.get('batchProcessing'):
            batch = result['batchProcessing']
            print(f"\n⚙️  배치 처리:")
            print(f"  • 등급 업데이트: {batch.get('gradeUpdates', 0)}명")
            print(f"  • 지급 계획 생성: {batch.get('paymentPlansCreated', 0)}건")

        print(f"{'='*60}\n")
        return True
    else:
        print(f"{'='*60}")
        print(f"❌ 업로드 실패: {response.status_code}")
        print(f"{'='*60}")
        try:
            error_data = response.json()
            print(f"오류: {error_data.get('error', '알 수 없는 오류')}")
        except:
            print(response.text[:500])
        print(f"{'='*60}\n")
        return False

def main():
    if len(sys.argv) < 3:
        print("사용법: python3 scripts/test/test_promotion_upload.py <파일경로> <귀속월>")
        print("\n예시:")
        print("  python3 scripts/test/test_promotion_upload.py test-data/test/7월_용역자명단_간단.xlsx 2025-07")
        print("  python3 scripts/test/test_promotion_upload.py test-data/test/8월_용역자명단_승급테스트.xlsx 2025-08")
        sys.exit(1)

    file_path = sys.argv[1]
    month_key = sys.argv[2]

    # 파일 존재 확인
    if not Path(file_path).exists():
        print(f"❌ 파일 없음: {file_path}")
        sys.exit(1)

    print("="*60)
    print(f"승급 시작일 테스트 - 엑셀 업로드")
    print("="*60)
    print(f"파일: {file_path}")
    print(f"귀속월: {month_key}")
    print("="*60 + "\n")

    # 로그인
    cookies = login_admin()

    # 엑셀 읽기
    users_data = read_excel_to_json(file_path)

    # 업로드
    file_name = Path(file_path).name
    success = upload_excel_data(cookies, users_data, month_key, file_name)

    if success:
        print("✅ 테스트 완료!")
    else:
        print("❌ 테스트 실패!")
        sys.exit(1)

if __name__ == "__main__":
    main()

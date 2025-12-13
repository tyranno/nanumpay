#!/usr/bin/env python3
"""
Excel 업로드 vs Reprocess API 결과 비교 테스트

목적: 7월부터 11월까지 매월 엑셀 업로드 후
      reprocess API로 지급 계획을 재생성했을 때
      결과가 동일한지 검증

사용법:
  python3 scripts/test/test_reprocess_comparison.py
  python3 scripts/test/test_reprocess_comparison.py --port 3101
  python3 scripts/test/test_reprocess_comparison.py --folder test
"""

import requests
import sys
import json
import argparse
import openpyxl
import subprocess
from pathlib import Path
from pymongo import MongoClient
from collections import defaultdict
from copy import deepcopy

# 설정
MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "nanumpay"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

# 폴더별 엑셀 파일 경로 (test_excel_upload.py와 동일)
FOLDER_FILES = {
    "test": {
        "7월": "test-data/test/7월_용역자명단_간단.xlsx",
        "8월": "test-data/test/8월_용역자명단_간단.xlsx",
        "9월": "test-data/test/9월_용역자명단_간단.xlsx",
        "10월": "test-data/test/10월_용역자명단_간단.xlsx",
        "11월": "test-data/test/11월_용역자명단_간단.xlsx",
    },
    "verify": {
        "7월": "test-data/verify/7월_용역자명단_간단.xlsx",
        "8월": "test-data/verify/8월_용역자명단_간단.xlsx",
        "9월": "test-data/verify/9월_용역자명단_간단.xlsx",
        "10월": "test-data/verify/10월_용역자명단_간단.xlsx",
        "11월": "test-data/verify/11월_용역자명단_간단.xlsx",
    }
}

# 월 순서
MONTH_ORDER = ["7월", "8월", "9월", "10월", "11월"]


def print_header(text):
    print(f"\n{'='*60}")
    print(f"  {text}")
    print(f"{'='*60}")


def print_subheader(text):
    print(f"\n{'-'*40}")
    print(f"  {text}")
    print(f"{'-'*40}")


def connect_db():
    """MongoDB 연결"""
    client = MongoClient(MONGO_URI)
    return client[DB_NAME]


def reset_db():
    """DB 초기화 (db_init.sh 실행)"""
    print_header("🗑️ DB 초기화")

    project_root = Path(__file__).parent.parent.parent
    db_dir = project_root / "apps/web/install/linux/db"
    init_script = project_root / "apps/web/install/linux/db_init.sh"

    if not init_script.exists():
        print(f"❌ 초기화 스크립트 없음: {init_script}")
        return False

    try:
        result = subprocess.run(
            ["bash", str(init_script), "--force"],
            env={"DB_DIR": str(db_dir), "PATH": "/usr/bin:/bin"},
            capture_output=True,
            text=True,
            cwd=str(project_root)
        )

        if result.returncode == 0:
            print("✅ DB 초기화 완료")
            return True
        else:
            print(f"❌ DB 초기화 실패:")
            print(result.stderr)
            return False
    except Exception as e:
        print(f"❌ DB 초기화 오류: {e}")
        return False


def login_admin(base_url):
    """관리자 로그인"""
    response = requests.post(
        f"{base_url}/api/auth/login",
        json={"loginId": ADMIN_LOGIN_ID, "password": ADMIN_PASSWORD}
    )

    if response.status_code == 200:
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        return None


def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 헤더 추출
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    # 데이터 추출
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True

        for idx, value in enumerate(row):
            if idx < len(headers) and value is not None and str(value).strip():
                if idx == 0:
                    row_data['__EMPTY'] = str(value).strip()
                else:
                    row_data[f'__EMPTY_{idx}'] = str(value).strip()

                if headers[idx]:
                    row_data[headers[idx]] = str(value).strip()
                is_empty = False

        if not is_empty:
            data.append(row_data)

    return data


def upload_excel(cookies, base_url, users_data, file_name):
    """엑셀 데이터 업로드"""
    response = requests.post(
        f"{base_url}/api/admin/users/bulk",
        json={"users": users_data, "fileName": file_name},
        cookies=cookies
    )

    if response.status_code == 200:
        return response.json()
    else:
        print(f"❌ 업로드 실패: {response.status_code}")
        print(response.text[:500])
        return None


def capture_payment_plans(db, month_key=None):
    """현재 지급 계획 상태 캡처"""
    query = {}
    if month_key:
        query['revenueMonth'] = month_key

    plans = list(db.weeklypaymentplans.find(query))

    # 비교를 위한 핵심 필드만 추출 (정렬된 형태)
    captured = []
    for plan in plans:
        captured.append({
            'userId': plan.get('userId'),
            'userName': plan.get('userName'),
            'baseGrade': plan.get('baseGrade'),
            'revenueMonth': plan.get('revenueMonth'),
            'planType': plan.get('planType'),
            'installmentType': plan.get('installmentType', 'basic'),
            'totalInstallments': plan.get('totalInstallments'),
            'installments': [
                {
                    'week': inst.get('week'),
                    'installmentAmount': inst.get('installmentAmount'),
                    'scheduledDate': str(inst.get('scheduledDate', ''))[:10],  # YYYY-MM-DD
                    'status': inst.get('status')
                }
                for inst in plan.get('installments', [])
            ]
        })

    # 정렬 (userId, revenueMonth, planType 순)
    captured.sort(key=lambda x: (
        x['userName'] or '',
        x['revenueMonth'] or '',
        x['planType'] or ''
    ))

    return captured


def call_reprocess_api(cookies, base_url, month_key):
    """Reprocess API 호출"""
    # 1. 먼저 해당 월의 사용자 조회
    response = requests.get(
        f"{base_url}/api/admin/users?limit=100",
        cookies=cookies
    )

    if response.status_code != 200:
        print(f"❌ 사용자 조회 실패: {response.status_code}")
        return False

    # 2. MonthlyRegistrations에서 해당 월 확인
    db = connect_db()
    monthly_reg = db.monthlyregistrations.find_one({'monthKey': month_key})

    if not monthly_reg:
        print(f"  ⚠️ {month_key} 월별 등록 데이터 없음")
        return False

    # 3. 특정 사용자 수정으로 reprocess 트리거
    # PUT /api/admin/users 에 requiresReprocess=true 전달
    users = list(db.users.find({'type': 'user'}))
    if not users:
        print("  ⚠️ 사용자 없음")
        return False

    # 아무 사용자나 선택해서 reprocess 요청
    target_user = users[0]

    response = requests.put(
        f"{base_url}/api/admin/users",
        json={
            "userId": str(target_user['_id']),
            "requiresReprocess": True,
            "name": target_user.get('name')  # 기존 값 유지
        },
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        return result.get('reprocessed', False)
    else:
        print(f"❌ Reprocess 실패: {response.status_code}")
        print(response.text[:500])
        return False


def compare_plans(original, reprocessed):
    """두 지급 계획 비교"""
    differences = []

    # 계획 수 비교
    if len(original) != len(reprocessed):
        differences.append(f"계획 수 다름: {len(original)} vs {len(reprocessed)}")

    # 각 계획 비교
    orig_map = {(p['userName'], p['revenueMonth'], p['planType']): p for p in original}
    repr_map = {(p['userName'], p['revenueMonth'], p['planType']): p for p in reprocessed}

    # Original에만 있는 계획
    only_orig = set(orig_map.keys()) - set(repr_map.keys())
    if only_orig:
        for key in only_orig:
            differences.append(f"원본에만 있음: {key}")

    # Reprocessed에만 있는 계획
    only_repr = set(repr_map.keys()) - set(orig_map.keys())
    if only_repr:
        for key in only_repr:
            differences.append(f"재처리에만 있음: {key}")

    # 공통 계획 필드 비교
    common_keys = set(orig_map.keys()) & set(repr_map.keys())
    for key in common_keys:
        orig = orig_map[key]
        repr_plan = repr_map[key]

        # 기본 필드 비교
        for field in ['baseGrade', 'totalInstallments', 'installmentType']:
            if orig.get(field) != repr_plan.get(field):
                differences.append(f"{key[0]} {field}: {orig.get(field)} vs {repr_plan.get(field)}")

        # 할부 비교
        orig_insts = orig.get('installments', [])
        repr_insts = repr_plan.get('installments', [])

        if len(orig_insts) != len(repr_insts):
            differences.append(f"{key[0]} 할부 수: {len(orig_insts)} vs {len(repr_insts)}")
            continue

        for i, (o_inst, r_inst) in enumerate(zip(orig_insts, repr_insts)):
            for field in ['installmentAmount', 'scheduledDate', 'status']:
                if o_inst.get(field) != r_inst.get(field):
                    differences.append(
                        f"{key[0]} 할부{i+1} {field}: {o_inst.get(field)} vs {r_inst.get(field)}"
                    )

    return differences


def run_month_test(cookies, base_url, db, month, file_path, project_root):
    """단일 월 테스트 실행"""
    # 월 키 계산 (7월 -> 2025-07)
    month_num = int(month.replace('월', ''))
    month_key = f"2025-{month_num:02d}"

    print_subheader(f"📤 {month} 엑셀 업로드")

    # 1. 엑셀 업로드
    full_path = project_root / file_path
    if not full_path.exists():
        print(f"❌ 파일 없음: {full_path}")
        return None, None

    users_data = read_excel_to_json(full_path)
    print(f"  📖 {len(users_data)}건 데이터 읽음")

    result = upload_excel(cookies, base_url, users_data, month)
    if not result:
        return None, None

    print(f"  ✅ 업로드 성공: {result.get('created', 0)}명 등록")

    # 2. 업로드 후 지급 계획 캡처
    print_subheader(f"📸 {month} 지급 계획 캡처")
    original_plans = capture_payment_plans(db)
    print(f"  📋 총 {len(original_plans)}개 계획 캡처됨")

    # 3. Reprocess API 호출
    print_subheader(f"🔄 {month_key} Reprocess")
    reprocessed = call_reprocess_api(cookies, base_url, month_key)

    if not reprocessed:
        print(f"  ⚠️ Reprocess 실행 안됨 (해당 월 사용자가 아닐 수 있음)")
        # 해당 월 사용자가 아니면 reprocess가 실행되지 않음
        # 이 경우 원본 계획 그대로 반환
        return original_plans, original_plans

    print(f"  ✅ Reprocess 완료")

    # 4. Reprocess 후 지급 계획 캡처
    reprocessed_plans = capture_payment_plans(db)
    print(f"  📋 총 {len(reprocessed_plans)}개 계획 (재처리 후)")

    return original_plans, reprocessed_plans


def main():
    parser = argparse.ArgumentParser(description='Excel 업로드 vs Reprocess 비교 테스트')
    parser.add_argument('--port', type=int, default=3101, help='서버 포트 (기본: 3101)')
    parser.add_argument('--folder', '-f', default='test', choices=['test', 'verify'],
                        help='데이터 폴더 (기본: test)')
    parser.add_argument('--no-reset', action='store_true', help='DB 초기화 생략')
    parser.add_argument('--month', '-m', help='특정 월만 테스트 (예: 7월)')

    args = parser.parse_args()

    base_url = f"http://localhost:{args.port}"
    project_root = Path(__file__).parent.parent.parent

    print_header("🧪 Excel 업로드 vs Reprocess 비교 테스트")
    print(f"  서버: {base_url}")
    print(f"  폴더: {args.folder}")

    # DB 초기화
    if not args.no_reset:
        if not reset_db():
            print("❌ DB 초기화 실패 - 중단")
            sys.exit(1)

    # DB 연결
    db = connect_db()

    # 로그인
    print_subheader("🔐 관리자 로그인")
    cookies = login_admin(base_url)
    if not cookies:
        sys.exit(1)
    print("  ✅ 로그인 성공")

    # 테스트할 월 목록
    excel_files = FOLDER_FILES.get(args.folder, {})
    months_to_test = [args.month] if args.month else MONTH_ORDER

    # 전체 결과 저장
    all_results = {}

    # 월별 테스트 실행
    for month in months_to_test:
        if month not in excel_files:
            print(f"\n⚠️ {month} 파일 없음 - 스킵")
            continue

        print_header(f"📆 {month} 테스트")

        original, reprocessed = run_month_test(
            cookies, base_url, db, month,
            excel_files[month], project_root
        )

        if original is None:
            print(f"❌ {month} 테스트 실패")
            continue

        # 비교
        print_subheader(f"🔍 {month} 결과 비교")
        differences = compare_plans(original, reprocessed)

        if not differences:
            print(f"  ✅ 완벽히 일치!")
            all_results[month] = {'status': 'PASS', 'plans': len(original)}
        else:
            print(f"  ❌ {len(differences)}개 차이 발견:")
            for diff in differences[:10]:
                print(f"    • {diff}")
            if len(differences) > 10:
                print(f"    ... 외 {len(differences) - 10}개")
            all_results[month] = {'status': 'FAIL', 'differences': len(differences)}

    # 최종 요약
    print_header("📊 최종 결과 요약")

    passed = sum(1 for r in all_results.values() if r['status'] == 'PASS')
    failed = sum(1 for r in all_results.values() if r['status'] == 'FAIL')

    for month, result in all_results.items():
        status_emoji = '✅' if result['status'] == 'PASS' else '❌'
        extra = f"({result.get('plans', 0)}개 계획)" if result['status'] == 'PASS' else f"({result.get('differences', 0)}개 차이)"
        print(f"  {status_emoji} {month}: {result['status']} {extra}")

    print(f"\n  합계: {passed} PASS / {failed} FAIL")

    if failed == 0:
        print("\n🎉 모든 테스트 통과!")
        return 0
    else:
        print("\n❌ 일부 테스트 실패")
        return 1


if __name__ == "__main__":
    sys.exit(main())

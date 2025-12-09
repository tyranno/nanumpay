#!/usr/bin/env python3
"""
등급별 지급 총액 조정 테스트 스크립트

테스트 시나리오:
1. 7월~11월 데이터 업로드
2. 11월 조정 (일반 모드) - F1:1만원, F2:2만원, F3:3만원, F4:4만원
3. 10월 조정 (force 모드) - 동일 금액

사용법:
  python3 scripts/test/test_grade_adjustment.py
"""

import requests
import sys
import json
import subprocess
import openpyxl
from pathlib import Path
from datetime import datetime

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

session = requests.Session()

def login_admin():
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")

    response = session.post(
        f"{BASE_URL}/api/auth/login",
        json={
            "loginId": ADMIN_LOGIN_ID,
            "password": ADMIN_PASSWORD
        }
    )

    if response.status_code == 200:
        print("✅ 관리자 로그인 성공")
        return True
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        print(response.text)
        return False

def get_payment_plans_from_db(month_key):
    """MongoDB에서 직접 지급 계획 조회"""
    query = f'''
    db.weeklypaymentplans.find({{revenueMonth: "{month_key}"}}).forEach(p => {{
        print(JSON.stringify({{
            _id: p._id.toString(),
            userId: p.userId,
            userName: p.userName,
            baseGrade: p.baseGrade,
            planType: p.planType,
            planStatus: p.planStatus,
            revenueMonth: p.revenueMonth,
            installments: p.installments.map(i => ({{
                installmentNumber: i.installmentNumber,
                installmentAmount: i.installmentAmount,
                status: i.status
            }}))
        }}));
    }});
    '''

    result = subprocess.run(
        ['mongosh', 'mongodb://localhost:27017/nanumpay', '--quiet', '--eval', query],
        capture_output=True,
        text=True
    )

    plans = []
    for line in result.stdout.strip().split('\n'):
        if line:
            try:
                plans.append(json.loads(line))
            except:
                pass

    return plans

def adjust_grade_payments(month_key, adjustments, force_update=False):
    """등급별 지급액 조정"""
    mode = "FORCE" if force_update else "일반"
    print(f"\n📊 등급별 지급액 조정: {month_key} [{mode} 모드]")

    # 금액 요약
    for grade, adj in adjustments.items():
        if adj.get('totalAmount'):
            per_inst = adj['totalAmount'] // 10
            print(f"   {grade}: 총 {adj['totalAmount']:,}원 → {per_inst:,}원/회")

    response = session.post(
        f"{BASE_URL}/api/admin/revenue/adjust-grade-payments",
        json={
            "monthKey": month_key,
            "adjustments": adjustments,
            "forceUpdate": force_update
        }
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ 조정 성공: {result.get('updatedPlans', 0)}개 계획 업데이트")
        return result
    else:
        print(f"❌ 조정 실패: {response.status_code}")
        print(response.text)
        return None

def print_payment_summary(plans, title="지급 계획 요약"):
    """지급 계획 요약 출력"""
    if not plans:
        print(f"\n⚠️ {title}: 데이터 없음")
        return

    print(f"\n{'='*70}")
    print(f"📋 {title}")
    print(f"{'='*70}")

    # 등급별 그룹화
    by_grade = {}
    for plan in plans:
        grade = plan.get('baseGrade', 'Unknown')
        if grade not in by_grade:
            by_grade[grade] = []
        by_grade[grade].append(plan)

    for grade in sorted(by_grade.keys()):
        grade_plans = by_grade[grade]
        print(f"\n[{grade}] {len(grade_plans)}개 계획")
        for plan in grade_plans:
            user_name = plan.get('userName', 'Unknown')
            plan_type = plan.get('planType', 'Unknown')
            installments = plan.get('installments', [])
            if installments:
                amount = installments[0].get('installmentAmount', 0)
            else:
                amount = 0
            status = plan.get('planStatus', 'Unknown')
            print(f"   • {user_name} ({plan_type}): {amount:,}원/회 [{status}]")

    print(f"{'='*70}")

def verify_adjustment(plans, expected_amounts, title="검증"):
    """조정 결과 검증"""
    print(f"\n🔍 {title}")
    all_passed = True
    checked = 0

    for grade, expected in expected_amounts.items():
        grade_plans = [p for p in plans if p.get('baseGrade') == grade]
        for plan in grade_plans:
            installments = plan.get('installments', [])
            if installments:
                actual = installments[0].get('installmentAmount', 0)
                status = plan.get('planStatus', '')
                checked += 1
                if actual == expected:
                    print(f"   ✅ {grade} {plan.get('userName')}: {actual:,}원 == {expected:,}원 [{status}]")
                else:
                    print(f"   ❌ {grade} {plan.get('userName')}: {actual:,}원 != {expected:,}원 [{status}]")
                    all_passed = False

    if checked == 0:
        print("   ⚠️ 검증할 데이터 없음")
        return False

    return all_passed

def upload_excel(month):
    """엑셀 데이터 업로드"""
    file_map = {
        "7월": "test-data/test/7월_용역자명단_간단.xlsx",
        "8월": "test-data/test/8월_용역자명단_간단.xlsx",
        "9월": "test-data/test/9월_용역자명단_간단.xlsx",
        "10월": "test-data/test/10월_용역자명단_간단.xlsx",
        "11월": "test-data/test/11월_용역자명단_간단.xlsx",
    }

    file_path = file_map.get(month)
    if not file_path:
        print(f"❌ {month} 파일 매핑 없음")
        return False

    full_path = Path(__file__).parent.parent.parent / file_path
    if not full_path.exists():
        print(f"❌ 파일 없음: {full_path}")
        return False

    wb = openpyxl.load_workbook(full_path)
    ws = wb.active

    # 첫 번째 행에서 헤더 읽기
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else '')

    # 데이터 행 읽기 (헤더명을 키로 사용)
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                if value is not None:
                    # 날짜는 ISO 포맷으로 변환
                    if headers[idx] == '날짜' and hasattr(value, 'strftime'):
                        row_data[headers[idx]] = value.strftime('%Y-%m-%d')
                    else:
                        row_data[headers[idx]] = str(value).strip() if value else ''

        # 성명이 있으면 데이터로 인정
        if row_data.get('성명'):
            data.append(row_data)

    if not data:
        print(f"❌ {month} 데이터가 비어있습니다")
        return False

    print(f"   📤 {month}: {len(data)}건...", end=" ")

    response = session.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": data, "fileName": f"{month}_용역자명단_간단.xlsx"}
    )

    if response.status_code == 200:
        result = response.json()
        created = result.get('created', 0)
        failed = result.get('failed', 0)
        print(f"✅ {created}명 등록, {failed}명 실패")
        if result.get('errors'):
            for err in result['errors'][:3]:
                print(f"      ⚠️ {err}")
        return True
    else:
        print(f"❌ 실패: {response.status_code}")
        try:
            print(f"      {response.json()}")
        except:
            print(f"      {response.text[:200]}")
        return False

def run_test():
    """메인 테스트 실행"""
    print("\n" + "="*70)
    print("🧪 등급별 지급 총액 조정 테스트")
    print("   - 11월: 일반 모드 (최근)")
    print("   - 10월: FORCE 모드 (2개월 전)")
    print("="*70)

    # 1. 로그인
    if not login_admin():
        return False

    # 2. 7월~11월 데이터 순차 업로드
    print("\n📁 Step 1: 7월~11월 데이터 업로드")
    for month in ["7월", "8월", "9월", "10월", "11월"]:
        if not upload_excel(month):
            print(f"⚠️ {month} 업로드 실패, 계속 진행...")

    # 3. 11월 초기 상태 확인
    print("\n" + "-"*70)
    print("📊 Step 2: 11월 초기 지급 계획 확인")
    plans_11 = get_payment_plans_from_db("2025-11")
    print_payment_summary(plans_11, "11월 초기 지급 계획")

    # 4. 10월 초기 상태 확인
    print("\n" + "-"*70)
    print("📊 Step 3: 10월 초기 지급 계획 확인")
    plans_10 = get_payment_plans_from_db("2025-10")
    print_payment_summary(plans_10, "10월 초기 지급 계획")

    # 조정할 금액 설정
    # F1: 1만원 → 10만원 총액 (1만원/회)
    # F2: 2만원 → 20만원 총액 (2만원/회)
    # F3: 3만원 → 30만원 총액 (3만원/회)
    # F4: 4만원 → 40만원 총액 (4만원/회)
    adjustments = {
        "F1": {"totalAmount": 100000},   # 10만원 총액 → 1만원/회
        "F2": {"totalAmount": 200000},   # 20만원 총액 → 2만원/회
        "F3": {"totalAmount": 300000},   # 30만원 총액 → 3만원/회
        "F4": {"totalAmount": 400000},   # 40만원 총액 → 4만원/회
    }

    expected = {
        "F1": 10000,
        "F2": 20000,
        "F3": 30000,
        "F4": 40000,
    }

    # 5. 11월 조정 (일반 모드)
    print("\n" + "-"*70)
    print("📝 Step 4: 11월 조정 [일반 모드]")
    result = adjust_grade_payments("2025-11", adjustments, force_update=False)

    if result:
        plans_11_after = get_payment_plans_from_db("2025-11")
        print_payment_summary(plans_11_after, "11월 조정 후")
        verify_adjustment(plans_11_after, expected, "11월 조정 결과 검증")

    # 6. 10월 조정 (FORCE 모드)
    print("\n" + "-"*70)
    print("📝 Step 5: 10월 조정 [FORCE 모드]")
    result = adjust_grade_payments("2025-10", adjustments, force_update=True)

    if result:
        plans_10_after = get_payment_plans_from_db("2025-10")
        print_payment_summary(plans_10_after, "10월 조정 후")
        verify_adjustment(plans_10_after, expected, "10월 조정 결과 검증 (FORCE)")

    print("\n" + "="*70)
    print("✅ 테스트 완료")
    print("="*70)

    return True

if __name__ == "__main__":
    try:
        success = run_test()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ 테스트 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

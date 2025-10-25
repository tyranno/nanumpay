#!/usr/bin/env python3
"""
승급 시 추가지급 중단 및 삭제 시 복원 테스트

흐름:
1. DB 초기화
2. 7~10월 등록
3. 이미영 추가지급 계획 확인 (10월 삭제 전)
4. 10월 삭제
5. 이미영 추가지급 계획 확인 (10월 삭제 후)
"""

import requests
import subprocess
import json

BASE_URL = "http://localhost:3100"
session = requests.Session()

def print_section(title):
    print(f"\n{'='*60}")
    print(f"📝 {title}")
    print('='*60)

def login():
    """관리자 로그인"""
    response = session.post(
        f"{BASE_URL}/api/auth/login",
        json={"loginId": "관리자", "password": "admin1234!!"}
    )
    if response.status_code == 200:
        print("✅ 로그인 성공")
        return True
    else:
        print(f"❌ 로그인 실패: {response.text}")
        return False

def init_db():
    """DB 초기화"""
    response = session.post(f"{BASE_URL}/api/admin/db/initialize")
    if response.status_code == 200:
        print("✅ DB 초기화 완료")
        return True
    else:
        print(f"❌ DB 초기화 실패: {response.text}")
        return False

def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True
        for idx, value in enumerate(row):
            if idx < len(headers):
                if value is not None and str(value).strip():
                    if idx == 0:
                        index_key = '__EMPTY'
                    else:
                        index_key = f'__EMPTY_{idx}'
                    row_data[index_key] = str(value).strip()
                    if headers[idx]:
                        row_data[headers[idx]] = str(value).strip()
                    is_empty = False
        if not is_empty:
            data.append(row_data)

    return data

def upload_month(month_name):
    """월별 용역자 등록 (JSON 방식)"""
    from pathlib import Path

    project_root = Path(__file__).parent.parent.parent
    file_path = project_root / f"test-data/test/{month_name}월_용역자명단_간단.xlsx"

    users_data = read_excel_to_json(file_path)

    response = session.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data}
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month_name}월 등록 성공: {result.get('created', 0)}명")
        return True
    else:
        result = response.json() if response.headers.get('content-type') == 'application/json' else {}
        print(f"❌ {month_name}월 등록 실패: {result.get('error', response.text)}")
        return False

def get_user_plans(user_name):
    """MongoDB에서 사용자의 모든 지급 계획 조회"""
    script = f'''
    const user = db.users.findOne({{ name: "{user_name}" }});
    if (!user) {{
        print(JSON.stringify({{ error: "User not found" }}));
    }} else {{
        const plans = db.weeklypaymentplans.find({{ userId: user._id.toString() }}).sort({{ baseGrade: 1, 추가지급단계: 1 }}).toArray();
        print(JSON.stringify({{
            grade: user.grade,
            plans: plans.map(p => ({{
                baseGrade: p.baseGrade,
                revenueMonth: p.revenueMonth,
                추가지급단계: p.추가지급단계 || 0,
                planStatus: p.planStatus,
                installmentType: p.installmentType,
                createdBy: p.createdBy,
                terminatedBy: p.terminatedBy,
                installments: p.installments ? p.installments.map(inst => ({{
                    week: inst.week,
                    scheduledDate: inst.scheduledDate,
                    installmentAmount: inst.installmentAmount,
                    status: inst.status
                }})) : []
            }}))
        }}));
    }}
    '''

    result = subprocess.run(
        ['mongosh', 'mongodb://localhost:27017/nanumpay', '--quiet', '--eval', script],
        capture_output=True,
        text=True
    )

    try:
        return json.loads(result.stdout.strip().split('\n')[-1])
    except:
        print(f"❌ 사용자 조회 실패: {result.stdout}")
        return None

def show_user_plans(user_name, title):
    """사용자 지급 계획 출력"""
    print(f"\n{'='*60}")
    print(f"📊 {title}")
    print('='*60)

    data = get_user_plans(user_name)
    if not data:
        print(f"❌ {user_name} 사용자를 찾을 수 없습니다.")
        return None

    print(f"\n{user_name}: {data['grade']} 등급, 지급 계획 {len(data['plans'])}건\n")

    for i, p in enumerate(data['plans'], 1):
        status_mark = ""
        if p['planStatus'] == 'terminated':
            status_mark = " [TERMINATED]"
        elif p['planStatus'] == 'completed':
            status_mark = " [COMPLETED]"

        term_mark = f" ← {p.get('terminatedBy', '')}" if p.get('terminatedBy') else ""

        print(f"  [{i}] {p['baseGrade']} 등급 - {p['revenueMonth']} 매출")
        print(f"      단계: {p['추가지급단계']} ({p.get('installmentType', 'basic')})")
        print(f"      상태: {p['planStatus']}{status_mark}{term_mark}")
        print(f"      생성: {p.get('createdBy', 'N/A')}")

        # 주차별 지급 일정 출력
        if p.get('installments'):
            print(f"      지급 일정 (10회):")
            for inst in p['installments']:
                date_str = inst['scheduledDate'][:10] if isinstance(inst['scheduledDate'], str) else str(inst['scheduledDate'])[:10]
                amount = f"{inst['installmentAmount']:,}원"
                status_str = inst['status']

                # terminated 표시
                status_display = status_str
                if status_str == 'terminated':
                    status_display = f"🚫 {status_str}"
                elif status_str == 'paid':
                    status_display = f"✅ {status_str}"
                elif status_str == 'pending':
                    status_display = f"⏳ {status_str}"

                print(f"        {inst['week']}회: {date_str} - {amount:>12} {status_display}")
        print()

    return data

def delete_month(month_key):
    """월별 데이터 삭제"""
    response = session.post(
        f"{BASE_URL}/api/admin/db/delete-monthly",
        json={"monthKey": month_key}
    )
    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month_key} 삭제 완료")
        print(f"   - 삭제된 사용자: {result.get('deletedUsers', 0)}명")
        print(f"   - 삭제된 지급 계획: {result.get('deletedPlans', 0)}건")
        if result.get('reprocessedMonth'):
            print(f"   - 재처리된 월: {result.get('reprocessedMonth')}")
        return result
    else:
        print(f"❌ {month_key} 삭제 실패: {response.text}")
        return None

def main():
    print_section("이미영 추가지급 계획 확인 테스트")

    # Step 1: 로그인
    if not login():
        return

    # Step 2: DB 초기화
    print_section("Step 1: DB 초기화")
    if not init_db():
        return

    # Step 3: 7~10월 등록
    print_section("Step 2: 7~10월 등록")
    months = ["7", "8", "9", "10"]

    for month_name in months:
        if not upload_month(month_name):
            return

    # Step 4: 이미영 추가지급 계획 확인 (10월 삭제 전)
    show_user_plans("이미영", "Step 3: 10월 삭제 전 - 이미영 지급 계획")

    # Step 5: 10월 삭제
    print_section("Step 4: 10월 삭제")
    delete_result = delete_month("2025-10")
    if not delete_result:
        return

    # Step 6: 이미영 추가지급 계획 확인 (10월 삭제 후)
    show_user_plans("이미영", "Step 5: 10월 삭제 후 - 이미영 지급 계획")

    # Step 7: 9월 삭제
    print_section("Step 6: 9월 삭제")
    delete_result = delete_month("2025-09")
    if not delete_result:
        return

    # Step 8: 이미영 추가지급 계획 확인 (9월 삭제 후 - canceled 복원 확인!)
    show_user_plans("이미영", "Step 7: 9월 삭제 후 - 이미영 지급 계획 (canceled 복원!)")

    # 검증
    print(f"\n{'='*60}")
    print("✅ 테스트 완료")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()

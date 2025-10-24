#!/usr/bin/env python3
"""
전체 Cascade 삭제 통합 테스트
DB 초기화 → 등록 → 삭제 → 고아 참조 확인을 한 번에 실행

사용법:
  python3 scripts/test/test_full_cascade.py
"""

import requests
import sys
import json
import subprocess
import openpyxl
from pathlib import Path

BASE_URL = "http://localhost:3100"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

def login_admin():
    """관리자 로그인"""
    print("🔐 관리자 로그인 중...")
    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"loginId": ADMIN_LOGIN_ID, "password": ADMIN_PASSWORD}
    )
    if response.status_code == 200:
        print(f"✅ 관리자 로그인 성공\n")
        return response.cookies
    else:
        print(f"❌ 로그인 실패: {response.status_code}")
        sys.exit(1)

def init_db(cookies):
    """DB 초기화 API 호출"""
    print("🗄️  DB 초기화 중...")
    response = requests.post(
        f"{BASE_URL}/api/admin/db/initialize",
        cookies=cookies
    )
    if response.status_code == 200:
        print(f"✅ DB 초기화 완료\n")
        return True
    else:
        print(f"❌ DB 초기화 실패: {response.status_code}\n")
        return False

def read_excel_to_json(file_path):
    """엑셀 파일을 JSON 배열로 변환"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 헤더 추출
    headers = []
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(None)

    # 데이터 추출
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        is_empty = True
        for idx, value in enumerate(row):
            if idx < len(headers):
                if value is not None and str(value).strip():
                    # 인덱스 기반 키 추가
                    if idx == 0:
                        index_key = '__EMPTY'
                    else:
                        index_key = f'__EMPTY_{idx}'
                    row_data[index_key] = str(value).strip()
                    # 헤더 키도 추가
                    if headers[idx]:
                        row_data[headers[idx]] = str(value).strip()
                    is_empty = False
        if not is_empty:
            data.append(row_data)

    return data

def upload_month(cookies, month_file):
    """월별 데이터 업로드"""
    print(f"📤 {month_file} 업로드 중...")

    # 엑셀 읽기
    project_root = Path(__file__).parent.parent.parent
    file_path = project_root / f"test-data/test/{month_file}"

    users_data = read_excel_to_json(file_path)

    # JSON으로 전송
    response = requests.post(
        f"{BASE_URL}/api/admin/users/bulk",
        json={"users": users_data},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month_file} 업로드 성공: {result.get('created', 0)}명 등록\n")
        return True
    else:
        print(f"❌ {month_file} 업로드 실패: {response.status_code}\n")
        return False

def delete_month(cookies, month_key):
    """월별 데이터 삭제"""
    print(f"🗑️  {month_key} 삭제 중...")
    response = requests.post(
        f"{BASE_URL}/api/admin/db/delete-monthly",
        json={"monthKey": month_key},
        cookies=cookies
    )

    if response.status_code == 200:
        result = response.json()
        print(f"✅ {month_key} 삭제 성공:")
        print(f"   - 용역자: {result.get('deletedUsers', 0)}명")
        print(f"   - 지급 계획: {result.get('deletedPlans', 0)}건\n")
        return True
    else:
        print(f"❌ {month_key} 삭제 실패: {response.status_code}")
        try:
            error = response.json()
            print(f"   오류: {error.get('error', '알 수 없음')}\n")
        except:
            print(f"   {response.text}\n")
        return False

def check_orphans():
    """고아 참조 확인"""
    print("🔍 고아 참조 확인 중...")

    script = """
    const users = db.users.find({}).toArray();
    let orphanCount = 0;
    const orphans = [];

    for (const user of users) {
        if (user.leftChildId) {
            const leftExists = db.users.findOne({_id: user.leftChildId});
            if (!leftExists) {
                orphanCount++;
                orphans.push({ user: user.name, field: 'leftChildId' });
            }
        }

        if (user.rightChildId) {
            const rightExists = db.users.findOne({_id: user.rightChildId});
            if (!rightExists) {
                orphanCount++;
                orphans.push({ user: user.name, field: 'rightChildId' });
            }
        }
    }

    print(JSON.stringify({ orphanCount, orphans, totalUsers: users.length }));
    """

    result = subprocess.run(
        ['mongosh', 'mongodb://localhost:27017/nanumpay', '--quiet', '--eval', script],
        capture_output=True,
        text=True
    )

    if result.returncode == 0:
        try:
            data = json.loads(result.stdout.strip().split('\n')[-1])
            print(f"   총 사용자: {data['totalUsers']}명")

            if data['orphanCount'] == 0:
                print(f"   ✅ 고아 참조 없음!\n")
                return True
            else:
                print(f"   ❌ 고아 참조 {data['orphanCount']}개 발견:")
                for orphan in data['orphans']:
                    print(f"      - {orphan['user']}.{orphan['field']}")
                print()
                return False
        except Exception as e:
            print(f"   ⚠️  파싱 실패: {e}\n")
            return False
    else:
        print(f"   ❌ MongoDB 쿼리 실패\n")
        return False

def main():
    print("\n" + "="*60)
    print("🚀 전체 Cascade 삭제 통합 테스트")
    print("="*60 + "\n")

    cookies = login_admin()

    # 1. DB 초기화
    if not init_db(cookies):
        sys.exit(1)

    # 2. 7~10월 등록
    print("📝 Step 1: 7~10월 데이터 등록")
    print("-" * 60)
    for month in ["7월", "8월", "9월", "10월"]:
        if not upload_month(cookies, f"{month}_용역자명단_간단.xlsx"):
            sys.exit(1)

    # 3. 고아 참조 확인 (등록 후)
    print("✓ Step 2: 등록 후 고아 참조 확인")
    print("-" * 60)
    if not check_orphans():
        print("❌ 등록 후에 고아 참조가 발견되었습니다!")
        sys.exit(1)

    # 4. 10월 → 7월 역순 삭제
    print("✓ Step 3: 10월 → 9월 → 8월 → 7월 역순 삭제")
    print("-" * 60)
    for month_key in ["2025-10", "2025-09", "2025-08", "2025-07"]:
        if not delete_month(cookies, month_key):
            sys.exit(1)

        # 각 삭제 후 고아 참조 확인
        if not check_orphans():
            print(f"❌ {month_key} 삭제 후 고아 참조 발견!")
            sys.exit(1)

    # 5. 최종 확인
    print("="*60)
    print("✅ 전체 통합 테스트 성공!")
    print("="*60)
    print("\n요약:")
    print("  ✅ DB 초기화")
    print("  ✅ 7~10월 등록 (고아 참조 없음)")
    print("  ✅ 10월 삭제 (고아 참조 없음)")
    print("  ✅ 9월 삭제 (고아 참조 없음)")
    print("  ✅ 8월 삭제 (고아 참조 없음)")
    print("  ✅ 7월 삭제 (고아 참조 없음)")
    print("\n🎉 Cascade 삭제 정상 작동!")
    print()

if __name__ == "__main__":
    main()

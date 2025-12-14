#!/usr/bin/env python3
"""
월 삭제 후 이전 월 재처리 비교 테스트

방법:
1. 7월 → 8월 → 9월 → 10월 → 11월 순차 업로드
2. 각 월 업로드 후 스냅샷 저장
3. 11월 삭제 → 10월 스냅샷과 비교
4. 10월 삭제 → 9월 스냅샷과 비교
5. 9월 삭제 → 8월 스냅샷과 비교
6. 8월 삭제 → 7월 스냅샷과 비교

사용법:
  python3 scripts/test/test_delete_and_compare.py
  python3 scripts/test/test_delete_and_compare.py --port 3101
"""

import requests
import sys
import json
import argparse
import openpyxl
import subprocess
from pathlib import Path
from pymongo import MongoClient
from copy import deepcopy

# 설정
MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "nanumpay"
ADMIN_LOGIN_ID = "관리자"
ADMIN_PASSWORD = "admin1234!!"

# 폴더별 엑셀 파일 경로
EXCEL_FILES = {
    "7월": "test-data/test/7월_용역자명단_간단.xlsx",
    "8월": "test-data/test/8월_용역자명단_간단.xlsx",
    "9월": "test-data/test/9월_용역자명단_간단.xlsx",
    "10월": "test-data/test/10월_용역자명단_간단.xlsx",
    "11월": "test-data/test/11월_용역자명단_간단.xlsx",
}

MONTH_ORDER = ["7월", "8월", "9월", "10월", "11월"]
MONTH_KEYS = {
    "7월": "2025-07",
    "8월": "2025-08",
    "9월": "2025-09",
    "10월": "2025-10",
    "11월": "2025-11",
}


def print_header(text):
    print(f"\n{'='*60}")
    print(f"  {text}")
    print(f"{'='*60}")


def print_subheader(text):
    print(f"\n{'-'*40}")
    print(f"  {text}")
    print(f"{'-'*40}")


def connect_db():
    client = MongoClient(MONGO_URI)
    return client[DB_NAME]


def reset_db():
    """DB 초기화"""
    print_header("🗑️ DB 초기화")

    project_root = Path(__file__).parent.parent.parent
    db_dir = project_root / "apps/web/install/linux/db"
    init_script = project_root / "apps/web/install/linux/db_init.sh"

    try:
        result = subprocess.run(
            ["bash", str(init_script), "--force"],
            env={"DB_DIR": str(db_dir), "PATH": "/usr/bin:/bin"},
            capture_output=True,
            text=True,
            cwd=str(project_root)
        )

        if result.returncode == 0:
            print("✅ DB 초기화 완료")
            return True
        else:
            print(f"❌ DB 초기화 실패: {result.stderr}")
            return False
    except Exception as e:
        print(f"❌ DB 초기화 오류: {e}")
        return False


def login_admin(base_url):
    response = requests.post(
        f"{base_url}/api/auth/login",
        json={"loginId": ADMIN_LOGIN_ID, "password": ADMIN_PASSWORD}
    )
    if response.status_code == 200:
        return response.cookies
    return None


def read_excel_to_json(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and value is not None and str(value).strip():
                row_data[f'__EMPTY_{idx}' if idx > 0 else '__EMPTY'] = str(value).strip()
                if headers[idx]:
                    row_data[headers[idx]] = str(value).strip()
        if row_data:
            data.append(row_data)

    return data


def upload_excel(cookies, base_url, users_data, file_name):
    response = requests.post(
        f"{base_url}/api/admin/users/bulk",
        json={"users": users_data, "fileName": file_name},
        cookies=cookies
    )
    if response.status_code == 200:
        return response.json()
    return None


def delete_month(cookies, base_url, month_key):
    """월 삭제 API 호출"""
    response = requests.post(
        f"{base_url}/api/admin/db/delete-monthly",
        json={"monthKey": month_key},
        cookies=cookies
    )
    if response.status_code == 200:
        return response.json()
    else:
        print(f"❌ 삭제 실패: {response.status_code}")
        print(response.text[:500])
        return None


def capture_snapshot(db):
    """현재 DB 상태 스냅샷 캡처"""
    snapshot = {
        'users': {},           # ⭐ dict로 변경 (키 기반 비교)
        'payment_plans': {},   # ⭐ dict로 변경 (키 기반 비교)
        'monthly_registrations': {}  # ⭐ dict로 변경
    }

    # Users (type='user'만) - ⭐ name을 키로 사용
    users = list(db.users.find({'type': 'user'}).sort([('name', 1)]))
    for user in users:
        user_key = user.get('name')
        # gradeHistory에서 승급 정보 추출
        grade_history = []
        for gh in user.get('gradeHistory', []):
            grade_history.append({
                'type': gh.get('type'),
                'revenueMonth': gh.get('revenueMonth'),
                'fromGrade': gh.get('fromGrade'),
                'toGrade': gh.get('toGrade'),
                'registrationDate': str(gh.get('registrationDate', ''))[:10] if gh.get('registrationDate') else None,
                'promotionDate': str(gh.get('promotionDate', ''))[:10] if gh.get('promotionDate') else None,
            })

        snapshot['users'][user_key] = {
            'name': user.get('name'),
            'grade': user.get('grade'),
            'status': user.get('status'),
            'leftChildId': str(user.get('leftChildId')) if user.get('leftChildId') else None,
            'rightChildId': str(user.get('rightChildId')) if user.get('rightChildId') else None,
            'gradeHistory': grade_history,  # ⭐ gradeHistory 추가
        }

    # Payment Plans - ⭐ 고유 키로 구분 (userName/revenueMonth/planType/installmentType)
    # ⭐ 정렬키에 installmentType 추가!
    plans = list(db.weeklypaymentplans.find().sort([
        ('userName', 1), ('revenueMonth', 1), ('planType', 1), ('installmentType', 1)
    ]))
    for plan in plans:
        # ⭐ 고유 키 생성 (installmentType 포함!)
        plan_key = f"{plan.get('userName')}/{plan.get('revenueMonth')}/{plan.get('planType')}/{plan.get('installmentType', 'basic')}"

        plan_data = {
            'userName': plan.get('userName'),
            'baseGrade': plan.get('baseGrade'),
            'revenueMonth': plan.get('revenueMonth'),
            'planType': plan.get('planType'),
            'planStatus': plan.get('planStatus'),
            'installmentType': plan.get('installmentType', 'basic'),
            'terminatedAt': str(plan.get('terminatedAt', ''))[:10] if plan.get('terminatedAt') else None,  # ⭐ 추가
            'terminationReason': plan.get('terminationReason'),  # ⭐ 추가
            'installments': []
        }
        for inst in plan.get('installments', []):
            plan_data['installments'].append({
                'week': inst.get('week'),
                'status': inst.get('status'),
                'installmentAmount': inst.get('installmentAmount'),
                'scheduledDate': str(inst.get('scheduledDate', ''))[:10]
            })
        snapshot['payment_plans'][plan_key] = plan_data

    # Monthly Registrations - ⭐ monthKey를 키로 사용
    regs = list(db.monthlyregistrations.find().sort('monthKey', 1))
    for reg in regs:
        month_key = reg.get('monthKey')
        snapshot['monthly_registrations'][month_key] = {
            'monthKey': month_key,
            'registrationCount': len(reg.get('registrations', [])),
            'totalRevenue': reg.get('totalRevenue', 0),
        }

    return snapshot


def compare_snapshots(expected, actual, label=""):
    """두 스냅샷 비교 - ⭐ 키 기반 비교로 변경"""
    differences = []

    # ========================================
    # Users 비교 (키 기반)
    # ========================================
    exp_user_keys = set(expected['users'].keys())
    act_user_keys = set(actual['users'].keys())

    missing_users = exp_user_keys - act_user_keys
    extra_users = act_user_keys - exp_user_keys

    if missing_users:
        for user_key in sorted(missing_users):
            differences.append(f"User 누락: {user_key}")

    if extra_users:
        for user_key in sorted(extra_users):
            differences.append(f"User 추가됨: {user_key}")

    common_users = exp_user_keys & act_user_keys
    for user_key in sorted(common_users):
        exp_user = expected['users'][user_key]
        act_user = actual['users'][user_key]

        for field in ['name', 'grade', 'status']:
            if exp_user.get(field) != act_user.get(field):
                differences.append(f"User {user_key} {field}: {exp_user.get(field)} vs {act_user.get(field)}")

        # ⭐ gradeHistory 비교
        exp_gh = exp_user.get('gradeHistory', [])
        act_gh = act_user.get('gradeHistory', [])
        if len(exp_gh) != len(act_gh):
            differences.append(f"User {user_key} gradeHistory 수: {len(exp_gh)} vs {len(act_gh)}")
        else:
            for i, (e, a) in enumerate(zip(exp_gh, act_gh)):
                for field in ['type', 'revenueMonth', 'fromGrade', 'toGrade', 'registrationDate', 'promotionDate']:
                    if e.get(field) != a.get(field):
                        differences.append(f"User {user_key} gradeHistory[{i}].{field}: {e.get(field)} vs {a.get(field)}")

    # ========================================
    # Payment Plans 비교 (키 기반) ⭐ 핵심 변경
    # ========================================
    exp_plan_keys = set(expected['payment_plans'].keys())
    act_plan_keys = set(actual['payment_plans'].keys())

    missing_plans = exp_plan_keys - act_plan_keys
    extra_plans = act_plan_keys - exp_plan_keys

    if missing_plans:
        for plan_key in sorted(missing_plans):
            differences.append(f"Plan 누락: {plan_key}")

    if extra_plans:
        for plan_key in sorted(extra_plans):
            differences.append(f"Plan 추가됨: {plan_key}")

    common_plans = exp_plan_keys & act_plan_keys
    for plan_key in sorted(common_plans):
        exp_plan = expected['payment_plans'][plan_key]
        act_plan = actual['payment_plans'][plan_key]

        # 필드 비교 (terminatedAt, terminationReason 포함)
        for field in ['baseGrade', 'planStatus', 'installmentType', 'terminatedAt', 'terminationReason']:
            if exp_plan.get(field) != act_plan.get(field):
                differences.append(f"Plan {plan_key} {field}: {exp_plan.get(field)} vs {act_plan.get(field)}")

        # Installments 비교
        exp_insts = exp_plan.get('installments', [])
        act_insts = act_plan.get('installments', [])

        if len(exp_insts) != len(act_insts):
            differences.append(f"Plan {plan_key} installments 수: {len(exp_insts)} vs {len(act_insts)}")
        else:
            for i, (exp_inst, act_inst) in enumerate(zip(exp_insts, act_insts)):
                # ⭐ week 필드도 포함
                for field in ['week', 'status', 'installmentAmount', 'scheduledDate']:
                    if exp_inst.get(field) != act_inst.get(field):
                        differences.append(f"Plan {plan_key} inst{i+1} {field}: {exp_inst.get(field)} vs {act_inst.get(field)}")

    # ========================================
    # Monthly Registrations 비교 (키 기반)
    # ========================================
    exp_reg_keys = set(expected['monthly_registrations'].keys())
    act_reg_keys = set(actual['monthly_registrations'].keys())

    missing_regs = exp_reg_keys - act_reg_keys
    extra_regs = act_reg_keys - exp_reg_keys

    if missing_regs:
        for reg_key in sorted(missing_regs):
            differences.append(f"MonthlyReg 누락: {reg_key}")

    if extra_regs:
        for reg_key in sorted(extra_regs):
            differences.append(f"MonthlyReg 추가됨: {reg_key}")

    common_regs = exp_reg_keys & act_reg_keys
    for reg_key in sorted(common_regs):
        exp_reg = expected['monthly_registrations'][reg_key]
        act_reg = actual['monthly_registrations'][reg_key]

        for field in ['registrationCount', 'totalRevenue']:
            if exp_reg.get(field) != act_reg.get(field):
                differences.append(f"MonthlyReg {reg_key} {field}: {exp_reg.get(field)} vs {act_reg.get(field)}")

    return differences


def main():
    parser = argparse.ArgumentParser(description='월 삭제 후 비교 테스트')
    parser.add_argument('--port', type=int, default=3101, help='서버 포트')
    args = parser.parse_args()

    base_url = f"http://localhost:{args.port}"
    project_root = Path(__file__).parent.parent.parent

    print_header("🧪 월 삭제 후 재처리 비교 테스트")
    print(f"  서버: {base_url}")

    # DB 초기화
    if not reset_db():
        sys.exit(1)

    db = connect_db()

    # 로그인
    print_subheader("🔐 관리자 로그인")
    cookies = login_admin(base_url)
    if not cookies:
        print("❌ 로그인 실패")
        sys.exit(1)
    print("  ✅ 로그인 성공")

    # ========================================
    # Phase 1: 순차 업로드 및 스냅샷 저장
    # ========================================
    print_header("📤 Phase 1: 순차 업로드 및 스냅샷 저장")

    snapshots = {}

    for month in MONTH_ORDER:
        print_subheader(f"📤 {month} 업로드")

        file_path = project_root / EXCEL_FILES[month]
        if not file_path.exists():
            print(f"❌ 파일 없음: {file_path}")
            continue

        users_data = read_excel_to_json(file_path)
        result = upload_excel(cookies, base_url, users_data, month)

        if result:
            print(f"  ✅ {result.get('created', 0)}명 등록")

            # 스냅샷 저장
            snapshots[month] = capture_snapshot(db)
            print(f"  📸 스냅샷 저장: Users={len(snapshots[month]['users'])}, Plans={len(snapshots[month]['payment_plans'])}")
        else:
            print(f"  ❌ 업로드 실패")

    # ========================================
    # Phase 2: 역순 삭제 및 비교
    # ========================================
    print_header("🗑️ Phase 2: 역순 삭제 및 비교")

    results = {}

    # 11월 → 10월 → 9월 → 8월 순으로 삭제
    delete_order = ["11월", "10월", "9월", "8월"]
    compare_targets = ["10월", "9월", "8월", "7월"]

    for delete_month, compare_month in zip(delete_order, compare_targets):
        print_subheader(f"🗑️ {delete_month} 삭제 → {compare_month} 스냅샷과 비교")

        month_key = MONTH_KEYS[delete_month]

        # 삭제 실행
        delete_result = delete_month_api(cookies, base_url, month_key)

        if delete_result:
            print(f"  ✅ 삭제 완료")
            print(f"     - 삭제된 용역자: {delete_result.get('deletedUsers', 0)}")
            print(f"     - 삭제된 계획: {delete_result.get('deletedPlans', 0)}")
            print(f"     - 재처리 월: {delete_result.get('reprocessedMonth', 'N/A')}")

            # 현재 상태 캡처
            current_snapshot = capture_snapshot(db)

            # 비교
            expected_snapshot = snapshots.get(compare_month)
            if expected_snapshot:
                differences = compare_snapshots(expected_snapshot, current_snapshot)

                if not differences:
                    print(f"  ✅ {compare_month} 스냅샷과 완벽히 일치!")
                    results[delete_month] = {'status': 'PASS'}
                else:
                    print(f"  ❌ {len(differences)}개 차이 발견:")
                    for diff in differences[:10]:
                        print(f"     • {diff}")
                    if len(differences) > 10:
                        print(f"     ... 외 {len(differences) - 10}개")
                    results[delete_month] = {'status': 'FAIL', 'differences': len(differences)}

                    # 상세 분석 출력 (⭐ dict 기반으로 변경)
                    print(f"\n  📋 상세 분석:")
                    print(f"     Expected plans ({len(expected_snapshot['payment_plans'])}):")

                    exp_plan_keys = set(expected_snapshot['payment_plans'].keys())
                    act_plan_keys = set(current_snapshot['payment_plans'].keys())

                    missing = exp_plan_keys - act_plan_keys
                    extra = act_plan_keys - exp_plan_keys

                    if missing:
                        print(f"     ❌ 누락된 계획 ({len(missing)}):")
                        for m in sorted(missing)[:10]:
                            print(f"        - {m}")
                    if extra:
                        print(f"     ➕ 추가된 계획 ({len(extra)}):")
                        for e in sorted(extra)[:10]:
                            print(f"        - {e}")

                    # ⭐ 공통 플랜 중 차이나는 것 상세 출력
                    common = exp_plan_keys & act_plan_keys
                    diff_details = []
                    for key in sorted(common):
                        exp_p = expected_snapshot['payment_plans'][key]
                        act_p = current_snapshot['payment_plans'][key]

                        for field in ['baseGrade', 'planStatus', 'installmentType', 'terminatedAt', 'terminationReason']:
                            if exp_p.get(field) != act_p.get(field):
                                diff_details.append(f"{key} {field}: {exp_p.get(field)} vs {act_p.get(field)}")

                        # installments 차이
                        for i, (e, a) in enumerate(zip(exp_p.get('installments', []), act_p.get('installments', []))):
                            for fld in ['status', 'installmentAmount', 'scheduledDate', 'week']:
                                if e.get(fld) != a.get(fld):
                                    diff_details.append(f"{key} inst{i+1}.{fld}: {e.get(fld)} vs {a.get(fld)}")

                    if diff_details:
                        print(f"     🔍 필드 차이 ({len(diff_details)}):")
                        for d in diff_details[:15]:
                            print(f"        - {d}")
                        if len(diff_details) > 15:
                            print(f"        ... 외 {len(diff_details) - 15}개")
            else:
                print(f"  ⚠️ {compare_month} 스냅샷 없음")
                results[delete_month] = {'status': 'SKIP'}
        else:
            print(f"  ❌ 삭제 실패")
            results[delete_month] = {'status': 'ERROR'}

    # ========================================
    # 최종 결과
    # ========================================
    print_header("📊 최종 결과")

    passed = sum(1 for r in results.values() if r['status'] == 'PASS')
    failed = sum(1 for r in results.values() if r['status'] == 'FAIL')

    for month, result in results.items():
        status_emoji = {'PASS': '✅', 'FAIL': '❌', 'SKIP': '⚠️', 'ERROR': '💥'}.get(result['status'], '❓')
        extra = f" ({result.get('differences', 0)}개 차이)" if result['status'] == 'FAIL' else ""
        print(f"  {status_emoji} {month} 삭제: {result['status']}{extra}")

    print(f"\n  합계: {passed} PASS / {failed} FAIL")

    if failed == 0:
        print("\n🎉 모든 테스트 통과!")
        return 0
    else:
        print("\n❌ 일부 테스트 실패")
        return 1


def delete_month_api(cookies, base_url, month_key):
    """월 삭제 API 호출"""
    response = requests.post(
        f"{base_url}/api/admin/db/delete-monthly",
        json={"monthKey": month_key},
        cookies=cookies
    )
    if response.status_code == 200:
        return response.json()
    else:
        print(f"  삭제 API 오류: {response.status_code}")
        try:
            print(f"  {response.json()}")
        except:
            print(f"  {response.text[:200]}")
        return None


if __name__ == "__main__":
    sys.exit(main())

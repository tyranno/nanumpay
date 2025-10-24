#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
테스트 데이터 생성 스크립트
- 7월 간단: 3명
- 7월 추가: 7명  
- 8월: 15명
- 9월: 45명
- 10월: 90명
이진 트리 구조 (각 부모 최대 2명 자식)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from collections import OrderedDict
import random

# ============================================
# 월별 데이터 정의 (ID와 이름 목록)
# 각 ID당 최대 8건 이하
# ============================================

# 7월 간단 (3명) - 루트 (고정)
JULY_SIMPLE = [
    ('사장님', ['사장님']),
    ('김영수', ['김영수']),
    ('이미영', ['이미영']),
]

# 7월 추가 (7명)
JULY_ADD = [
    ('박철수', ['박철수', '박철수2', '박철수3', '박철수4']),  # 4명
    ('강민지', ['강민지']),  # 1명
    ('이순신', ['이순신']),  # 1명
    ('김유신', ['김유신']),  # 1명
]

# 8월 (15명)
AUGUST = [
    ('한예진', ['한예진', '한예진2']),  # 2명
    ('황민정', ['황민정', '황민정2']),  # 2명  
    ('신윤복', ['신윤복', '신윤복2']),  # 2명
    ('정선', ['정선', '정선2']),  # 2명
    ('박지원', ['박지원', '박지원2', '박지원3']),  # 3명
    ('윤서준', ['윤서준', '윤서준2', '윤서준3', '윤서준4']),  # 4명
]

# 9월 (45명) - 10개 ID
SEPTEMBER = [
    ('최민수', ['최민수', '최민수2', '최민수3', '최민수4', '최민수5', '최민수6']),  # 6명
    ('송가인', ['송가인', '송가인2', '송가인3', '송가인4', '송가인5']),  # 5명
    ('류현진', ['류현진', '류현진2', '류현진3', '류현진4', '류현진5']),  # 5명
    ('안정환', ['안정환', '안정환2', '안정환3', '안정환4', '안정환5']),  # 5명
    ('배용준', ['배용준', '배용준2', '배용준3', '배용준4']),  # 4명
    ('전지현', ['전지현', '전지현2', '전지현3', '전지현4']),  # 4명
    ('고소영', ['고소영', '고소영2', '고소영3', '고소영4']),  # 4명
    ('김하늘', ['김하늘', '김하늘2', '김하늘3', '김하늘4']),  # 4명
    ('이나영', ['이나영', '이나영2', '이나영3', '이나영4']),  # 4명
    ('한지민', ['한지민', '한지민2', '한지민3', '한지민4']),  # 4명
]

# 10월 (90명) - 30개 ID, 다양한 인원수
OCTOBER = [
    ('김태희', ['김태희']),  # 1명
    ('정우성', ['정우성', '정우성2', '정우성3', '정우성4', '정우성5']),  # 5명
    ('손예진', ['손예진', '손예진2']),  # 2명
    ('현빈', ['현빈', '현빈2', '현빈3', '현빈4']),  # 4명
    ('이병헌', ['이병헌']),  # 1명
    ('송혜교', ['송혜교', '송혜교2', '송혜교3']),  # 3명
    ('강동원', ['강동원', '강동원2', '강동원3', '강동원4', '강동원5', '강동원6']),  # 6명
    ('전도연', ['전도연', '전도연2']),  # 2명
    ('하정우', ['하정우', '하정우2', '하정우3', '하정우4']),  # 4명
    ('김혜수', ['김혜수']),  # 1명
    ('조인성', ['조인성', '조인성2', '조인성3']),  # 3명
    ('송강호', ['송강호', '송강호2', '송강호3', '송강호4', '송강호5']),  # 5명
    ('이정재', ['이정재', '이정재2']),  # 2명
    ('공유', ['공유']),  # 1명
    ('김수현', ['김수현', '김수현2', '김수현3', '김수현4', '김수현5', '김수현6', '김수현7']),  # 7명
    ('박서준', ['박서준', '박서준2', '박서준3']),  # 3명
    ('이민호', ['이민호', '이민호2']),  # 2명
    ('김지원', ['김지원', '김지원2', '김지원3', '김지원4']),  # 4명
    ('박보영', ['박보영']),  # 1명
    ('한효주', ['한효주', '한효주2', '한효주3', '한효주4', '한효주5']),  # 5명
    ('수지', ['수지', '수지2']),  # 2명
    ('아이유', ['아이유', '아이유2', '아이유3']),  # 3명
    ('김고은', ['김고은', '김고은2', '김고은3', '김고은4']),  # 4명
    ('박신혜', ['박신혜']),  # 1명
    ('전소민', ['전소민', '전소민2', '전소민3', '전소민4', '전소민5', '전소민6']),  # 6명
    ('김다미', ['김다미', '김다미2']),  # 2명
    ('이성경', ['이성경', '이성경2', '이성경3']),  # 3명
    ('한소희', ['한소희', '한소희2', '한소희3', '한소희4']),  # 4명
    ('정호연', ['정호연']),  # 1명
    ('김세정', ['김세정', '김세정2', '김세정3', '김세정4', '김세정5']),  # 5명
]


# ============================================
# 전역 변수 - 자식 수 관리
# ============================================

childData = OrderedDict()  # 각 이름의 자식 수를 저장 (순서 보장)


# ============================================
# 헬퍼 함수들
# ============================================

def initialize_child_data():
    """childData 초기화 - 모든 데이터의 이름을 0으로 초기화"""
    global childData
    childData = OrderedDict()
    
    # 7월 간단 초기화 (사장님은 자식 2명으로 고정)
    childData['사장님'] = 2  # 김영수, 이미영을 자식으로 가짐
    childData['김영수'] = 0
    childData['이미영'] = 0
    
    # 7월 추가 데이터 초기화
    for _, names in JULY_ADD:
        for name in names:
            childData[name] = 0
    
    # 8월 데이터 초기화
    for _, names in AUGUST:
        for name in names:
            childData[name] = 0
    
    # 9월 데이터 초기화
    for _, names in SEPTEMBER:
        for name in names:
            childData[name] = 0
    
    # 10월 데이터 초기화
    for _, names in OCTOBER:
        for name in names:
            childData[name] = 0
    
    print(f"✅ childData 초기화 완료: 총 {len(childData)}명")


def find_parent_for_name(current_name):
    """
    현재 이름보다 앞선 사람 중에서 랜덤하게 부모 선택
    단, 자식 2개인 노드가 25%를 넘지 않도록 제어
    """
    global childData
    
    # 현재 이름의 위치 찾기
    all_names = list(childData.keys())
    
    # 현재 이름이 childData에 없으면 None 반환
    if current_name not in all_names:
        return None
    
    current_index = all_names.index(current_name)
    
    # 앞에 있는 사람들 중 자식이 2명 미만인 후보 모으기
    candidates_0 = []  # 자식 0명
    candidates_1 = []  # 자식 1명
    
    for i in range(current_index):
        parent_name = all_names[i]
        if childData[parent_name] == 0:
            candidates_0.append(parent_name)
        elif childData[parent_name] == 1:
            candidates_1.append(parent_name)
    
    if not candidates_0 and not candidates_1:
        return None
    
    # 현재 통계 계산 (자식이 있는 노드들만 카운트)
    count_2 = sum(1 for c in childData.values() if c == 2)
    count_1 = sum(1 for c in childData.values() if c == 1)
    total_with_children = count_2 + count_1
    
    ratio_2 = count_2 / total_with_children if total_with_children > 0 else 0
    
    # 자식 2개인 노드 비율에 따라 선택 전략 조정
    if ratio_2 >= 0.30:  # 이미 30% 이상이면 자식 0명만 선택
        if candidates_0:
            candidates = candidates_0
        else:
            candidates = candidates_1  # 어쩔 수 없이 1명 선택
    else:  # 25% 미만이면 자식 1명인 후보 우선 (2개로 만들기)
        if candidates_1 and random.random() < 0.7:  # 70% 확률로 자식 1명 선택
            candidates = candidates_1
        elif candidates_0:
            candidates = candidates_0
        else:
            candidates = candidates_1
    
    # 랜덤하게 부모 선택
    parent = random.choice(candidates)
    
    childData[parent] += 1
    return parent


def generate_person_data(name, counter, account_id=None):
    """개인 데이터 자동 생성 - 판매인은 childData 기반으로 할당"""
    banks = ['KB국민은행', '신한은행', '우리은행', '하나은행', '농협은행', '기업은행', 'NH농협은행', '카카오뱅크', '토스뱅크']

    # account_id가 없으면 name을 사용
    if account_id is None:
        account_id = name

    # 판매인 찾기
    if name == '사장님':
        salesperson = '본인'  # 사장님은 루트이므로 본인
    elif name in ['김영수', '이미영']:
        salesperson = '사장님'  # 7월 간단의 김영수, 이미영은 사장님 자식으로 고정
    else:
        salesperson = find_parent_for_name(name)
        if not salesperson:
            salesperson = '본인'  # 부모를 찾을 수 없으면 본인
    
    # 전화번호 생성 (010-XXXX-XXXX 형식)
    phone_middle = 1000 + (counter * 7) % 9000
    phone_last = 1000 + (counter * 13) % 9000
    
    # 주민번호 생성 (YYMMDD-1XXXXXX 또는 YYMMDD-2XXXXXX)
    year = 70 + (counter % 30)  # 70~99 (1970~1999년생)
    month = 1 + (counter % 12)
    day = 1 + (counter % 28)
    gender = 1 + (counter % 2)  # 1 또는 2
    random_digits = 1000000 + (counter * 123456) % 9000000
    id_number = f'{year:02d}{month:02d}{day:02d}-{gender}{random_digits:07d}'[:14]  # 14자리로 제한
    
    # 계좌번호 생성 (은행별 실제 형식)
    bank = banks[counter % len(banks)]
    if '국민' in bank:
        # KB국민은행: XXXXXX-XX-XXXXXX (6-2-6)
        account = f'{100000 + counter:06d}-{10 + counter % 90:02d}-{100000 + counter * 7 % 900000:06d}'
    elif '신한' in bank:
        # 신한은행: XXX-XXX-XXXXXX (3-3-6)
        account = f'{100 + counter % 900:03d}-{100 + counter % 900:03d}-{100000 + counter * 3 % 900000:06d}'
    elif '우리' in bank:
        # 우리은행: XXXX-XXX-XXXXXX (4-3-6)
        account = f'{1000 + counter % 9000:04d}-{100 + counter % 900:03d}-{100000 + counter * 5 % 900000:06d}'
    elif '하나' in bank:
        # 하나은행: XXX-XXXXXX-XXXXX (3-6-5)
        account = f'{100 + counter % 900:03d}-{100000 + counter % 900000:06d}-{10000 + counter * 11 % 90000:05d}'
    elif '농협' in bank:
        # 농협: XXX-XXXX-XXXX-XX (3-4-4-2)
        account = f'{100 + counter % 900:03d}-{1000 + counter % 9000:04d}-{1000 + counter * 7 % 9000:04d}-{10 + counter % 90:02d}'
    elif '기업' in bank:
        # 기업은행: XXX-XXXXXX-XX-XXX (3-6-2-3)
        account = f'{100 + counter % 900:03d}-{100000 + counter % 900000:06d}-{10 + counter % 90:02d}-{100 + counter % 900:03d}'
    elif '카카오' in bank:
        # 카카오뱅크: XXXX-XX-XXXXXXX (4-2-7)
        account = f'{3333:04d}-{1 + counter % 99:02d}-{1000000 + counter * 17 % 9000000:07d}'
    elif '토스' in bank:
        # 토스뱅크: XXXX-XXXX-XXXX (4-4-4)
        account = f'{1000 + counter % 9000:04d}-{1000 + counter * 3 % 9000:04d}-{1000 + counter * 7 % 9000:04d}'
    else:
        # 기타: XXX-XXXXXX-XXXXX
        account = f'{100 + counter % 900:03d}-{100000 + counter % 900000:06d}-{10000 + counter % 90000:05d}'
    
    # 지사 전화번호
    branch_middle = 2000 + (counter * 11) % 8000
    branch_last = 3000 + (counter * 17) % 7000
    
    # 설계사 전화번호
    designer_middle = 5000 + (counter * 23) % 5000
    designer_last = 4000 + (counter * 29) % 6000
    
    # 설계사 이름 (좀 더 다양하게)
    designer_suffixes = ['설계', '매니저', '팀장', '과장', '대리', '주임', '실장', '부장']
    designer_name = f'{name[0]}{designer_suffixes[counter % len(designer_suffixes)]}'

    # ⭐ 보험 상품 및 회사 정보
    insurance_products = [
        '무배당 KB 5.10.10 플랜',
        '(무) 하나로 든든한 내일',
        '무배당 삼성화재 든든한 미래',
        '무배당 메리츠 안심케어',
        '무배당 DB손해보험 프리미엄',
        '무배당 현대해상 평생케어',
        '무배당 AIA 건강플랜',
        '무배당 푸르덴셜 가족사랑',
        '무배당 교보생명 안심보장',
        '무배당 한화생명 든든플랜'
    ]

    insurance_companies = [
        'KB손해보험',
        '하나생명',
        '삼성화재',
        '메리츠화재',
        'DB손해보험',
        '현대해상',
        'AIA생명',
        '푸르덴셜생명',
        '교보생명',
        '한화생명'
    ]

    insurance_product = insurance_products[counter % len(insurance_products)]
    insurance_company = insurance_companies[counter % len(insurance_companies)]

    # ⭐ 지사명
    branch_offices = [
        '서울본사',
        '강남지사',
        '강북지사',
        '분당지사',
        '인천지사',
        '경기지사',
        '부산지사',
        '대구지사',
        '광주지사',
        '대전지사'
    ]
    branch_office = branch_offices[counter % len(branch_offices)]

    return {
        'user_id': account_id,  # ⭐ 계정ID (같은 계정은 같은 ID)
        'name': name,
        'phone': f'010-{phone_middle:04d}-{phone_last:04d}',
        'idNumber': id_number,
        'bank': bank,
        'account': account,
        'salesperson': salesperson,
        'branch': f'010-{branch_middle:04d}-{branch_last:04d}',
        'designer': designer_name,
        'designer_phone': f'010-{designer_middle:04d}-{designer_last:04d}',
        'insurance_product': insurance_product,  # ⭐ 보험상품명
        'insurance_company': insurance_company,  # ⭐ 보험회사
        'branch_office': branch_office           # ⭐ 지사명
    }


def create_excel(filename, data_list, registration_month):
    """Excel 파일 생성"""
    import os
    from datetime import datetime, timedelta
    
    # 저장 경로 설정
    save_dir = os.path.join(os.path.dirname(__file__), '..', 'test-data')
    os.makedirs(save_dir, exist_ok=True)
    filepath = os.path.join(save_dir, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '용역자 명단'
    
    headers = ['순번', '날짜', 'ID', '성명', '연락처', '주민번호', '은행', '계좌번호', '판매인', '연락처', '설계사', '연락처', '보험상품명', '보험회사', '지사']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # 헤더 설정
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 해당 월의 시작 날짜 계산
    year, month = map(int, registration_month.split('-'))
    start_date = datetime(year, month, 1)
    
    # 데이터 입력
    for idx, person in enumerate(data_list):
        row_num = idx + 2
        # 날짜는 해당 월의 1일부터 순차적으로 (같은 날에 여러 명 등록 가능)
        date_offset = idx // 3  # 3명씩 같은 날
        current_date = start_date + timedelta(days=date_offset)
        date_str = current_date.strftime('%Y-%m-%d')
        
        ws.cell(row=row_num, column=1).value = idx + 1  # 순번 (1부터 시작)
        ws.cell(row=row_num, column=2).value = date_str  # 날짜
        ws.cell(row=row_num, column=3).value = person['user_id']  # ID
        ws.cell(row=row_num, column=4).value = person['name']  # 성명
        ws.cell(row=row_num, column=5).value = person['phone']  # 연락처
        ws.cell(row=row_num, column=6).value = person['idNumber']  # 주민번호
        ws.cell(row=row_num, column=7).value = person['bank']  # 은행
        ws.cell(row=row_num, column=8).value = person['account']  # 계좌번호
        ws.cell(row=row_num, column=9).value = person['salesperson']  # 판매인
        ws.cell(row=row_num, column=10).value = person['branch']  # 연락처 (판매인)
        ws.cell(row=row_num, column=11).value = person['designer']  # 설계사
        ws.cell(row=row_num, column=12).value = person['designer_phone']  # 연락처 (설계사)
        ws.cell(row=row_num, column=13).value = person['insurance_product']  # ⭐ 보험상품명
        ws.cell(row=row_num, column=14).value = person['insurance_company']  # ⭐ 보험회사
        ws.cell(row=row_num, column=15).value = person['branch_office']  # ⭐ 지사명
    
    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filepath)
    print(f'✅ {filepath} 생성 완료 ({len(data_list)}명)')


def process_month_data(month_data, month_name):
    """월별 데이터 처리 - 튜플 데이터를 엑셀용 데이터로 변환"""
    result = []
    counter = len(childData)  # 현재까지 등록된 사람 수를 카운터로 사용

    for account_id, names in month_data:  # ⭐ account_id는 원본 계정명
        for name in names:
            # childData에 이미 있는 이름이면 건너뜀 (중복 방지)
            if name not in childData:
                childData[name] = 0  # 새로운 이름을 childData에 추가

            person_data = generate_person_data(name, counter, account_id)  # ⭐ account_id 전달
            result.append(person_data)
            counter += 1

    return result


def main():
    print('🚀 테스트 데이터 생성 시작...\n')
    
    # 1단계: childData 초기화
    print('📊 1단계: childData 초기화')
    initialize_child_data()
    print()
    
    # 데이터 수 확인
    july_simple_count = sum(len(names) for _, names in JULY_SIMPLE)
    july_add_count = sum(len(names) for _, names in JULY_ADD)
    august_count = sum(len(names) for _, names in AUGUST)
    september_count = sum(len(names) for _, names in SEPTEMBER)
    october_count = sum(len(names) for _, names in OCTOBER)
    
    print('📊 데이터 정의 확인:')
    print(f'  7월 간단: {july_simple_count}명 (ID: {len(JULY_SIMPLE)}개)')
    print(f'  7월 추가: {july_add_count}명 (ID: {len(JULY_ADD)}개)')
    print(f'  8월: {august_count}명 (ID: {len(AUGUST)}개)')
    print(f'  9월: {september_count}명 (ID: {len(SEPTEMBER)}개)')
    print(f'  10월: {october_count}명 (ID: {len(OCTOBER)}개)')
    print()
    
    # 7월 간단은 기존 파일 사용 (생성하지 않음)
    print('📊 7월 간단은 기존 파일 사용')
    
    # 나머지 월별 데이터 처리
    print('\n📊 나머지 월별 데이터 처리')
    
    july_add_data = process_month_data(JULY_ADD, '7월 추가')
    print(f'✅ 7월 추가: {len(july_add_data)}명')
    
    august_data = process_month_data(AUGUST, '8월')
    print(f'✅ 8월: {len(august_data)}명')
    
    september_data = process_month_data(SEPTEMBER, '9월')
    print(f'✅ 9월: {len(september_data)}명')
    
    october_data = process_month_data(OCTOBER, '10월')
    print(f'✅ 10월: {len(october_data)}명')
    
    # Excel 파일 생성 (7월 간단은 기존 파일 사용)
    print('\n📊 Excel 파일 생성 (7월 간단은 기존 파일 사용)')
    create_excel('7월_용역자명단_추가.xlsx', july_add_data, '2025-07')
    create_excel('8월_용역자명단_간단.xlsx', august_data, '2025-08')
    create_excel('9월_용역자명단_간단.xlsx', september_data, '2025-09')
    create_excel('10월_용역자명단_간단.xlsx', october_data, '2025-10')
    
    print('\n📊 전체 요약:')
    print(f'  7월 간단: 3명 (기존 파일 사용)')
    print(f'  7월 추가: {len(july_add_data)}명 (누적 10명)')  
    print(f'  8월: {len(august_data)}명 (누적 {10 + len(august_data)}명)')
    print(f'  9월: {len(september_data)}명 (누적 {10 + len(august_data) + len(september_data)}명)')
    print(f'  10월: {len(october_data)}명 (누적 {10 + len(august_data) + len(september_data) + len(october_data)}명)')
    print('\n✅ 모든 테스트 데이터 생성 완료!')


if __name__ == '__main__':
    main()
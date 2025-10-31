# -*- coding: utf-8 -*-
"""
계산.xlsx 파일에서 등급 항목을 자동으로 채워넣는 스크립트

규칙:
1. B9 셀 (승급자): 등급, 승 형식으로 표시 (예: F1->F2 라면 "2,승")
2. B11 셀 (미승급자): "1" 로 표시
3. B13 셀 (추가지급 인원): 등급, 차수 형식으로 표시 (예: F1, 1차 라면 "1,추1")
"""

from __future__ import print_function
import openpyxl
from openpyxl import load_workbook
import re

def extract_grade_number(grade_str):
    """등급 문자열에서 숫자 추출 (예: F1 -> 1, F2 -> 2)"""
    match = re.search(r'F(\d+)', grade_str)
    if match:
        return match.group(1)
    return None

def extract_round_number(round_str):
    """차수 문자열에서 숫자 추출 (예: 1차 -> 1, 2차 -> 2)"""
    match = re.search(r'(\d+)차', round_str)
    if match:
        return match.group(1)
    return None

def process_promotion(value):
    """승급자 처리: F1->F2 라면 2,승 반환"""
    if '->' in str(value):
        parts = value.split('->')
        if len(parts) == 2:
            target_grade = parts[1].strip()
            grade_num = extract_grade_number(target_grade)
            if grade_num:
                return "{},승".format(grade_num)
    return None

def process_additional_payment(grade_value, round_value):
    """추가지급 인원 처리: F1, 1차 라면 1,추1 반환"""
    grade_num = extract_grade_number(str(grade_value))
    round_num = extract_round_number(str(round_value))

    if grade_num and round_num:
        return "{},추{}".format(grade_num, round_num)
    return None

def main():
    # 파일 로드
    import sys
    import os
    reload(sys)
    sys.setdefaultencoding('utf-8')

    # 현재 디렉토리의 파일명을 직접 확인
    file_name = None
    for f in os.listdir(u'.'):
        if f.endswith('.xlsx') and u'계산' in f and not f.startswith('~$'):
            file_name = f
            break

    if not file_name:
        print(u"계산.xlsx 파일을 찾을 수 없습니다.")
        return

    print(u"파일 로드 중: {}".format(file_name))
    wb = load_workbook(file_name)
    ws = wb.active

    # B9 셀의 시작 행 찾기 (승급자)
    promotion_start_row = None
    non_promotion_start_row = None
    additional_payment_start_row = None

    # B 열을 스캔하여 각 섹션의 시작 행 찾기
    for row in range(1, ws.max_row + 1):
        cell_value = ws['B{}'.format(row)].value
        if cell_value == u'승급자':
            promotion_start_row = row
        elif cell_value == u'미승급자':
            non_promotion_start_row = row
        elif cell_value == u'추가지급 인원':
            additional_payment_start_row = row

    print(u"승급자 섹션: B{}".format(promotion_start_row))
    print(u"미승급자 섹션: B{}".format(non_promotion_start_row))
    print(u"추가지급 인원 섹션: B{}".format(additional_payment_start_row))

    # 등급 열이 어디에 있는지 확인 (첫 번째 행에서 '등급' 찾기)
    grade_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=promotion_start_row, column=col).value == u'등급':
            grade_col = col
            break

    if not grade_col:
        print(u"등급 열을 찾을 수 없습니다.")
        return

    grade_col_letter = openpyxl.utils.get_column_letter(grade_col)
    print(u"등급 열: {}".format(grade_col_letter))

    # 1. 승급자 처리
    if promotion_start_row:
        current_row = promotion_start_row + 1
        while current_row < (non_promotion_start_row if non_promotion_start_row else ws.max_row):
            name_cell = ws['B{}'.format(current_row)].value
            if not name_cell or str(name_cell).strip() == '':
                break

            # 이름 옆 셀에서 승급 정보 확인 (예: C 열에 F1->F2 같은 정보가 있다고 가정)
            # 실제 구조에 맞게 조정 필요
            for check_col in range(3, 10):  # C부터 I까지 체크
                check_value = ws.cell(row=current_row, column=check_col).value
                if check_value and '->' in str(check_value):
                    result = process_promotion(check_value)
                    if result:
                        ws.cell(row=current_row, column=grade_col).value = result
                        print(u"B{} ({}): {}".format(current_row, name_cell, result))
                        break

            current_row += 1

    # 2. 미승급자 처리
    if non_promotion_start_row:
        current_row = non_promotion_start_row + 1
        while current_row < (additional_payment_start_row if additional_payment_start_row else ws.max_row):
            name_cell = ws['B{}'.format(current_row)].value
            if not name_cell or str(name_cell).strip() == '':
                break

            ws.cell(row=current_row, column=grade_col).value = "1"
            print(u"B{} ({}): 1".format(current_row, name_cell))
            current_row += 1

    # 3. 추가지급 인원 처리
    if additional_payment_start_row:
        current_row = additional_payment_start_row + 1
        while current_row <= ws.max_row:
            name_cell = ws['B{}'.format(current_row)].value
            if not name_cell or str(name_cell).strip() == '':
                break

            # 등급과 차수 정보 찾기
            grade_value = None
            round_value = None

            for check_col in range(3, 10):
                check_value = ws.cell(row=current_row, column=check_col).value
                if check_value:
                    check_str = str(check_value)
                    if 'F' in check_str and extract_grade_number(check_str):
                        grade_value = check_str
                    if u'차' in check_str and extract_round_number(check_str):
                        round_value = check_str

            if grade_value and round_value:
                result = process_additional_payment(grade_value, round_value)
                if result:
                    ws.cell(row=current_row, column=grade_col).value = result
                    print(u"B{} ({}): {}".format(current_row, name_cell, result))

            current_row += 1

    # 저장
    output_file = u'계산_updated.xlsx'
    wb.save(output_file)
    print(u"\n완료! 결과 파일: {}".format(output_file))

if __name__ == "__main__":
    main()

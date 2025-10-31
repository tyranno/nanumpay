# -*- coding: utf-8 -*-
from __future__ import print_function
import openpyxl
from openpyxl import load_workbook
import sys
import os

reload(sys)
sys.setdefaultencoding('utf-8')

# 현재 디렉토리의 파일명을 직접 확인
file_name = None
for f in os.listdir(u'.'):
    if f.endswith('.xlsx') and u'계산' in f and not f.startswith('~$'):
        file_name = f
        break

if not file_name:
    print(u"계산.xlsx 파일을 찾을 수 없습니다.")
    sys.exit(1)

print(u"파일 로드 중: {}".format(file_name))
wb = load_workbook(file_name)
ws = wb.active

print(u"\n파일 구조 확인:")
print(u"최대 행: {}".format(ws.max_row))
print(u"최대 열: {}".format(ws.max_column))

print(u"\n첫 30행의 B열 내용:")
for row in range(1, min(31, ws.max_row + 1)):
    cell_value = ws['B{}'.format(row)].value
    if cell_value:
        print(u"B{}: {}".format(row, cell_value))

print(u"\n첫 행의 모든 열:")
for col in range(1, min(20, ws.max_column + 1)):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(u"{}: {}".format(col_letter, cell_value))

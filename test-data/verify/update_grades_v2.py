# -*- coding: utf-8 -*-
"""
계산.xlsx 파일에서 등급 항목을 자동으로 채워넣는 스크립트

규칙:
1. 승급자 (B9 행): 등급, 승 형식으로 표시 (예: F1->F2 라면 "2,승")
2. 미승급자 (B11 행): "1" 로 표시
3. 추가지급 인원 (B13 행): 등급, 차수 형식으로 표시 (예: F1, 1차 라면 "1,추1")
"""

from __future__ import print_function
import openpyxl
from openpyxl import load_workbook
import re
import sys
import os

def extract_grade_number(grade_str):
    """등급 문자열에서 숫자 추출 (예: F1 -> 1, F2 -> 2)"""
    match = re.search(r'F(\d+)', grade_str)
    if match:
        return match.group(1)
    return None

def extract_round_number(round_str):
    """차수 문자열에서 숫자 추출 (예: 1차 -> 1, 2차 -> 2)"""
    match = re.search(r'(\d+)', round_str)
    if match:
        return match.group(1)
    return None

def parse_name_info(text):
    """텍스트에서 이름과 정보 추출"""
    # 예: "강동원2(F1→F2)" -> ("강동원2", "F1→F2")
    # 예: "김태희(F1)" -> ("김태희", "F1")
    # 예: "최민수2(F1,1차)" -> ("최민수2", "F1,1차")
    match = re.match(r'([^(]+)\(([^)]+)\)', text.strip())
    if match:
        return match.group(1), match.group(2)
    return None, None

def main():
    reload(sys)
    sys.setdefaultencoding('utf-8')

    # 현재 디렉토리의 파일명을 직접 확인
    file_name = None
    for f in os.listdir(u'.'):
        if f.endswith('.xlsx') and u'계산' in f and not f.startswith('~$'):
            file_name = f
            break

    if not file_name:
        print(u"계산.xlsx 파일을 찾을 수 없습니다.")
        return

    print(u"파일 로드 중: {}".format(file_name))
    wb = load_workbook(file_name)
    ws = wb.active

    # 이름과 등급 정보를 저장할 딕셔너리
    grade_info = {}

    # 1. B9 행에서 승급자 정보 추출
    promotion_data = ws['B9'].value
    if promotion_data:
        parts = promotion_data.split(u'→ ')[1] if u'→ ' in promotion_data else promotion_data
        for item in parts.split(', '):
            name, info = parse_name_info(item)
            if name and info and u'→' in info:
                # F1→F2 형식
                target_grade = info.split(u'→')[1].strip()
                grade_num = extract_grade_number(target_grade)
                if grade_num:
                    grade_info[name] = u"{},승".format(grade_num)
                    print(u"승급자 - {}: {}".format(name, grade_info[name]))

    # 2. B11 행에서 미승급자 정보 추출
    non_promotion_data = ws['B11'].value
    if non_promotion_data:
        parts = non_promotion_data.split(u'→ ')[1] if u'→ ' in non_promotion_data else non_promotion_data
        for item in parts.split(', '):
            name, info = parse_name_info(item)
            if name:
                grade_info[name] = u"1"
                print(u"미승급자 - {}: {}".format(name, grade_info[name]))

    # 3. B13 행에서 추가지급 인원 정보 추출
    additional_data = ws['B13'].value
    if additional_data:
        parts = additional_data.split(u'→ ')[1] if u'→ ' in additional_data else additional_data
        for item in parts.split(', '):
            name, info = parse_name_info(item)
            if name and info and ',' in info:
                # F1,1차 형식
                grade_part = info.split(',')[0].strip()
                round_part = info.split(',')[1].strip()
                grade_num = extract_grade_number(grade_part)
                round_num = extract_round_number(round_part)
                if grade_num and round_num:
                    grade_info[name] = u"{},추{}".format(grade_num, round_num)
                    print(u"추가지급 - {}: {}".format(name, grade_info[name]))

    # 4. B22행부터 시작되는 이름 목록 처리
    # 첫 번째로 "이름"이 있는 행을 찾기
    name_start_row = None
    for row in range(1, ws.max_row + 1):
        if ws['B{}'.format(row)].value == u'이름':
            name_start_row = row
            break

    if not name_start_row:
        print(u"이름 열을 찾을 수 없습니다.")
        return

    print(u"\n이름 목록 시작 행: B{}".format(name_start_row))

    # 등급 열 찾기 (첫 번째 행에서 '등급' 찾기)
    grade_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=name_start_row, column=col).value == u'등급':
            grade_col = col
            break

    if not grade_col:
        print(u"등급 열을 찾을 수 없습니다.")
        return

    grade_col_letter = openpyxl.utils.get_column_letter(grade_col)
    print(u"등급 열: {}".format(grade_col_letter))

    # 이름 목록에 등급 정보 채우기
    current_row = name_start_row + 1
    updated_count = 0
    while current_row <= ws.max_row:
        name = ws['B{}'.format(current_row)].value
        if not name or str(name).strip() == '':
            break

        if name in grade_info:
            ws.cell(row=current_row, column=grade_col).value = grade_info[name]
            print(u"B{} ({}) <- {}".format(current_row, name, grade_info[name]))
            updated_count += 1
        else:
            print(u"B{} ({}) - 정보 없음".format(current_row, name))

        current_row += 1

    # 저장
    output_file = u'계산_updated.xlsx'
    wb.save(output_file)
    print(u"\n완료! {} 개 항목 업데이트".format(updated_count))
    print(u"결과 파일: {}".format(output_file))

if __name__ == "__main__":
    main()

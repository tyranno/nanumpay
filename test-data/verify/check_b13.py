# -*- coding: utf-8 -*-
from __future__ import print_function
import openpyxl
from openpyxl import load_workbook
import sys
import os

reload(sys)
sys.setdefaultencoding('utf-8')

file_name = None
for f in os.listdir(u'.'):
    if f.endswith('.xlsx') and u'계산' in f and not f.startswith('~$'):
        file_name = f
        break

wb = load_workbook(file_name)
ws = wb.active

b13_value = ws['B13'].value
print(u"B13 전체 내용:")
print(b13_value)
print(u"\n")

# 화살표로 분리
if u'→ ' in b13_value:
    parts = b13_value.split(u'→ ')[1]
    print(u"화살표 이후:")
    print(parts)
    print(u"\n")

    # 콤마로 분리
    items = parts.split(', ')
    print(u"첫 5개 항목:")
    for i, item in enumerate(items[:5]):
        print(u"{}: '{}'".format(i, item))
